#!/usr/bin/env python3
"""
MIoT 批量创建工具（通用版）
- 自动查询产品服务列表
- 从 Excel 读取属性、方法、事件定义
- 自动识别属性类型（bool/数值/枚举）
- 批量调用 API 创建属性、方法、事件
"""

import argparse
import json
import os
import sys
import time

import requests
from openpyxl import load_workbook

# ─── API ──────────────────────────────────────────────────────
BASE = "https://iot.mi.com"
CREATE_PROP_API   = f"{BASE}/cgi-std/post/api/v1/functionDefine/addInstanceProperty"
CREATE_ACTION_API = f"{BASE}/cgi-std/post/api/v1/functionDefine/addInstanceAction"
CREATE_EVENT_API  = f"{BASE}/cgi-std/post/api/v1/functionDefine/addInstanceEvent"
QUERY_PROPS_API   = f"{BASE}/cgi-std/api/v1/functionDefine/getInstanceProperties"
QUERY_SERVICES_API = f"{BASE}/cgi-std/api/v1/functionDefine/getInstanceServices"

HEADERS = {
    "accept": "application/json, text/plain, */*",
    "content-type": "application/json",
    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/146.0.0.0 Safari/537.36",
    "origin": BASE,
    "referer": f"{BASE}/",
}


# ─── Excel 读取 ──────────────────────────────────────────────

def read_config(ws) -> dict:
    """读取公共配置 Sheet"""
    cfg = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key, val = row[0], row[1]
        if key and val is not None:
            cfg[str(key).strip()] = str(val).strip() if val else ""
    return cfg


def read_sheet_items(ws, name_col_idx: int = 2) -> list[dict]:
    """
    通用读取 Sheet（属性/方法/事件），返回列表
    name_col_idx: name 列的索引（0-based），默认2（第3列=属性name列）
    第一行是表头，第二行是说明，数据从第三行开始
    """
    headers = None
    items = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if i == 2:
            continue
        # 空行跳过 - 用 name 列判断
        name_val = row[name_col_idx] if len(row) > name_col_idx else None
        if not name_val:
            continue
        d = {}
        for j, h in enumerate(headers):
            val = row[j] if j < len(row) else None
            d[h] = val
        items.append(d)
    return items


def read_properties(ws) -> list[dict]:
    """读取属性定义 Sheet，返回属性列表（name在第3列，index=2）"""
    return read_sheet_items(ws, name_col_idx=2)


def read_actions(ws) -> list[dict]:
    """读取方法定义 Sheet，返回方法列表（name在第1列，index=0）"""
    return read_sheet_items(ws, name_col_idx=0)


def read_events(ws) -> list[dict]:
    """读取事件定义 Sheet，返回事件列表（name在第1列，index=0）"""
    return read_sheet_items(ws, name_col_idx=0)


# ─── API 请求 ─────────────────────────────────────────────────

def build_cookies(config: dict) -> dict:
    """构建请求 Cookie"""
    return {
        "serviceToken": config.get("serviceToken", ""),
        "userId": config.get("userId", ""),
        "xiaomiiot_ph": config.get("xiaomiiot_ph", ""),
    }


def build_query_params(config: dict, **extra) -> dict:
    """构建 Query 参数"""
    params = {
        "userId": config.get("userId", ""),
        "xiaomiiot_ph": config.get("xiaomiiot_ph", ""),
        "pdId": config.get("pdId", ""),
    }
    params.update(extra)
    return params


def _safe_request(method: str, url: str, *, max_retries: int = 3,
                  retry_delay: float = 1.0, **kwargs) -> requests.Response:
    """带重试的 HTTP 请求，网络错误自动重试"""
    last_exc = None
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.request(method, url, timeout=15, **kwargs)
            return resp
        except (ConnectionResetError, ConnectionError, OSError) as e:
            last_exc = e
            if attempt < max_retries:
                print(f"  ⚠️ 请求失败(第{attempt}次)，{retry_delay}s后重试: {e}")
                time.sleep(retry_delay)
    raise last_exc  # type: ignore


def query_services(config: dict) -> list[dict]:
    """查询产品的所有服务列表"""
    params = build_query_params(
        config,
        model=config.get("model", ""),
        connectType=config.get("connectType", "16"),
        language=config.get("language", "zh_cn"),
        version=config.get("version", "1"),
        status=config.get("status", "0"),
    )
    resp = _safe_request("GET", QUERY_SERVICES_API, params=params,
                         headers=HEADERS, cookies=build_cookies(config))
    data = resp.json()
    if data.get("status") != 200:
        print(f"❌ 查询服务列表失败: {data}")
        return []
    return data.get("result", [])


def query_properties(siid: int, service_type: str, config: dict) -> list[dict]:
    """查询某个服务下的已有属性"""
    params = build_query_params(config)
    params.update({
        "version": config.get("version", "1"),
        "status": config.get("status", "0"),
        "siid": str(siid),
        "serviceType": service_type,
        "model": config.get("model", ""),
        "connectType": config.get("connectType", "16"),
        "language": config.get("language", "zh_cn"),
    })
    resp = _safe_request("GET", QUERY_PROPS_API, params=params,
                         headers=HEADERS, cookies=build_cookies(config))
    data = resp.json()
    if data.get("status") != 200:
        return []
    return data.get("result", [])


def create_property(body: dict, config: dict) -> dict:
    """调用创建属性接口"""
    params = build_query_params(config)
    resp = _safe_request("POST", CREATE_PROP_API, params=params, json=body,
                         headers=HEADERS, cookies=build_cookies(config))
    return resp.json()


def create_action(body: dict, config: dict) -> dict:
    """调用创建方法接口"""
    params = build_query_params(config)
    resp = _safe_request("POST", CREATE_ACTION_API, params=params, json=body,
                         headers=HEADERS, cookies=build_cookies(config))
    return resp.json()


def create_event(body: dict, config: dict) -> dict:
    """调用创建事件接口"""
    params = build_query_params(config)
    resp = _safe_request("POST", CREATE_EVENT_API, params=params, json=body,
                         headers=HEADERS, cookies=build_cookies(config))
    return resp.json()


# ─── 属性类型识别 & 请求体构造 ───────────────────────────────

def detect_value_type(fmt: str, prop: dict) -> str:
    """自动识别值类型: bool_range / number / enum / string"""
    vt = str(prop.get("value_type") or "").strip().lower()
    if vt in ("bool_range", "number", "enum", "string"):
        return vt

    # 自动判断
    if fmt == "bool":
        return "bool_range"
    if fmt == "string":
        return "string"

    # 有枚举值定义 → enum
    vl = str(prop.get("value_list") or "").strip()
    if vl:
        return "enum"

    # 有数值范围 → number
    vr_min = prop.get("value_range_min")
    if vr_min is not None and str(vr_min).strip():
        return "number"

    # 默认
    return "number"


def parse_value_list(raw: str) -> list:
    """解析枚举值: '0:关闭,1:开启,2:待机' → [{value:0,description:'关闭'}, ...]"""
    if not raw or not str(raw).strip():
        return []
    result = []
    for item in str(raw).strip().split(","):
        item = item.strip()
        if ":" in item:
            val, desc = item.split(":", 1)
            result.append({"value": int(val.strip()), "description": desc.strip()})
        elif item.isdigit():
            result.append({"value": int(item), "description": f"值{item}"})
    return result


def parse_access(raw: str) -> list:
    """解析权限: 'read,write,notify' → ['read','write','notify']"""
    if not raw:
        return ["read", "write", "notify"]
    return [a.strip() for a in str(raw).split(",") if a.strip()]


def parse_bool(raw) -> bool:
    """解析布尔值"""
    if isinstance(raw, bool):
        return raw
    s = str(raw).strip().lower()
    return s in ("true", "1", "yes")


def build_request_body(prop: dict, config: dict, service_info: dict = None) -> dict:
    """构造创建属性的请求体"""

    def safe_int(val, default=0):
        """安全转换为 int，空字符串返回 default"""
        if val is None or str(val).strip() == "":
            return default
        try:
            return int(val)
        except (ValueError, TypeError):
            return default

    fmt = str(prop.get("format", "bool")).strip()
    vtype = detect_value_type(fmt, prop)

    # valueList & valueRange
    if vtype == "bool_range":
        value_list = []
        value_range = []
    elif vtype == "enum":
        value_list = parse_value_list(prop.get("value_list", ""))
        value_range = []
    elif vtype == "string":
        value_list = []
        value_range = []
    else:  # number
        value_list = []
        vr_min = prop.get("value_range_min")
        vr_max = prop.get("value_range_max")
        vr_step = prop.get("value_range_step")
        value_range = [
            int(vr_min) if vr_min is not None and str(vr_min).strip() else 0,
            int(vr_max) if vr_max is not None and str(vr_max).strip() else 65535,
            int(vr_step) if vr_step is not None and str(vr_step).strip() else 1,
        ]

    # 服务信息 - 优先用属性行内值，其次用查到的服务信息，最后用公共配置
    siid = prop.get("siid")
    if siid is None or str(siid).strip() == "":
        if service_info:
            siid = service_info.get("siid")
        else:
            siid = 0
    # 确保 siid 是 int（服务返回的 siid 可能是字符串）
    siid = safe_int(siid, 0)

    service_type = prop.get("service_type") or prop.get("serviceType", "")
    if not str(service_type).strip() and service_info:
        service_type = service_info.get("type", "")

    body = {
        "language": config.get("language", "zh_cn") or "zh_cn",
        "description": str(prop.get("description", "")),
        "format": fmt,
        "valueList": value_list,
        "valueRange": value_range,
        "access": parse_access(prop.get("access") or config.get("access", "read,write,notify")),
        "source": safe_int(prop.get("source") or config.get("source"), 4),
        "name": str(prop.get("name", "")),
        "gattAccess": parse_access(prop.get("access") or config.get("access", "read,write,notify")),
        "version": safe_int(config.get("version"), 1),
        "status": safe_int(config.get("status"), 0),
        "siid": safe_int(siid, 0),
        "serviceType": str(service_type),
        "model": str(prop.get("model") or config.get("model", "")),
        "connectType": safe_int(prop.get("connectType") or config.get("connectType"), 16),
        "standard": parse_bool(prop.get("standard") or config.get("standard", False)),
        "specCategory": None,
        "pdId": safe_int(prop.get("pdId") or config.get("pdId"), 0),
    }
    return body


def build_action_request_body(item: dict, config: dict, service_info: dict = None) -> dict:
    """构造创建方法的请求体"""

    def safe_int(val, default=0):
        if val is None or str(val).strip() == "":
            return default
        try:
            return int(val)
        except (ValueError, TypeError):
            return default

    # 服务信息
    siid = item.get("siid")
    if siid is None or str(siid).strip() == "":
        siid = service_info.get("siid") if service_info else 0
    siid = safe_int(siid, 0)

    service_type = item.get("service_type") or item.get("serviceType", "")
    if not str(service_type).strip() and service_info:
        service_type = service_info.get("type", "")

    name = str(item.get("name", ""))
    normalization_desc = str(item.get("normalizationDesc", "") or name)

    return {
        "version": safe_int(config.get("version"), 1),
        "status": safe_int(config.get("status"), 0),
        "siid": siid,
        "serviceType": str(service_type),
        "model": str(config.get("model", "")),
        "connectType": safe_int(config.get("connectType"), 16),
        "language": config.get("language", "zh_cn") or "zh_cn",
        "standard": parse_bool(item.get("standard") or config.get("standard", False)),
        "name": name,
        "normalizationDesc": normalization_desc,
        "description": str(item.get("description", "")),
        "pdId": safe_int(config.get("pdId"), 0),
    }


def build_event_request_body(item: dict, config: dict, service_info: dict = None) -> dict:
    """构造创建事件的请求体（与方法结构几乎一致）"""

    def safe_int(val, default=0):
        if val is None or str(val).strip() == "":
            return default
        try:
            return int(val)
        except (ValueError, TypeError):
            return default

    siid = item.get("siid")
    if siid is None or str(siid).strip() == "":
        siid = service_info.get("siid") if service_info else 0
    siid = safe_int(siid, 0)

    service_type = item.get("service_type") or item.get("serviceType", "")
    if not str(service_type).strip() and service_info:
        service_type = service_info.get("type", "")

    name = str(item.get("name", ""))
    normalization_desc = str(item.get("normalizationDesc", "") or name)

    return {
        "version": safe_int(config.get("version"), 1),
        "status": safe_int(config.get("status"), 0),
        "siid": siid,
        "serviceType": str(service_type),
        "model": str(config.get("model", "")),
        "connectType": safe_int(config.get("connectType"), 16),
        "language": config.get("language", "zh_cn") or "zh_cn",
        "standard": parse_bool(item.get("standard") or config.get("standard", False)),
        "name": name,
        "normalizationDesc": normalization_desc,
        "description": str(item.get("description", "")),
        "pdId": safe_int(config.get("pdId"), 0),
    }


# ─── 服务匹配 ────────────────────────────────────────────────

def match_service(prop: dict, services: list[dict]) -> dict | None:
    """
    根据属性行的 service_name / service_desc / siid 匹配服务
    优先级:
      1. service_name + service_desc 同时精确匹配（区分同名服务）
      2. service_desc 精确匹配
      3. service_name + service_desc 模糊匹配
      4. service_name 精确匹配
      5. service_desc 包含匹配
      6. siid 直接匹配（仅当以上都无法匹配时兜底）
    注意：siid 优先级最低，因为换产品后 siid 会变化
    """
    prop_siid = prop.get("siid")
    prop_sname = str(prop.get("service_name") or "").strip()
    prop_sdesc = str(prop.get("service_desc") or "").strip()

    # 1. service_name + service_desc 同时精确匹配
    if prop_sname and prop_sdesc:
        for svc in services:
            if svc.get("name") == prop_sname and svc.get("description") == prop_sdesc:
                return svc

    # 2. service_desc 精确匹配
    if prop_sdesc:
        for svc in services:
            if svc.get("description") == prop_sdesc:
                return svc

    # 3. service_name + service_desc 模糊匹配
    if prop_sname and prop_sdesc:
        for svc in services:
            svc_name = svc.get("name", "")
            svc_desc = svc.get("description", "")
            if prop_sname in svc_name and prop_sdesc in svc_desc:
                return svc

    # 4. service_name 精确匹配（仅当 desc 为空时）
    if prop_sname:
        for svc in services:
            if svc.get("name") == prop_sname:
                return svc

    # 5. service_desc 包含匹配
    if prop_sdesc:
        for svc in services:
            if prop_sdesc in svc.get("description", ""):
                return svc

    # 6. siid 直接匹配（兜底，仅在名称/描述都无法匹配时使用）
    if prop_siid is not None and str(prop_siid).strip():
        try:
            target_siid = int(prop_siid)
            for svc in services:
                svc_siid = svc.get("siid")
                if svc_siid is not None and int(svc_siid) == target_siid:
                    return svc
        except (ValueError, TypeError):
            pass

    return None


# ─── 批量执行 ────────────────────────────────────────────────

def batch_create(tasks: list[dict], create_fn, config: dict,
                 label: str, id_field: str, delay: float) -> tuple[int, int, list]:
    """
    通用批量创建函数
    create_fn: 创建函数 (body, config) -> response
    label: 类型标签（属性/方法/事件）
    id_field: 创建成功返回的 ID 字段名（piid/aiid/eiid）
    """
    success = 0
    failed = 0
    results = []

    print(f"\n🚀 开始创建 {len(tasks)} 条{label}...\n")
    for t in tasks:
        print(f"  [{t['index']}] {t['name']} ({t['desc']}) → siid={t['siid']} ... ", end="", flush=True)
        try:
            resp = create_fn(t["body"], config)
            status = resp.get("status")
            result_val = resp.get("result")
            if status == 200:
                new_id = result_val
                print(f"✅ 成功 ({id_field}={new_id})")
                success += 1
                results.append({"name": t["name"], "status": "success", id_field: new_id, "siid": t["siid"]})
            else:
                msg = resp.get("message", resp.get("msg", json.dumps(resp, ensure_ascii=False)))
                print(f"❌ 失败 ({msg})")
                failed += 1
                results.append({"name": t["name"], "status": "failed", "error": msg, "siid": t["siid"]})
        except Exception as e:
            print(f"❌ 异常 ({e})")
            failed += 1
            results.append({"name": t["name"], "status": "error", "error": str(e), "siid": t["siid"]})

        time.sleep(delay)

    print(f"\n{'='*50}")
    print(f"📊 {label}创建完成: 成功 {success} / 失败 {failed} / 共 {len(tasks)}")
    if failed > 0:
        print(f"\n❌ 失败列表:")
        for r in results:
            if r["status"] != "success":
                print(f"  - {r['name']} (siid={r['siid']}): {r.get('error', '未知错误')}")

    return success, failed, results


# ─── 主流程 ──────────────────────────────────────────────────

def print_services(services: list[dict]):
    """打印服务列表"""
    print(f"\n📋 产品服务列表（共 {len(services)} 个）:")
    print("-" * 80)
    print(f"{'siid':>4} | {'name':<22} | {'description':<20} | type")
    print("-" * 80)
    for svc in services:
        siid = svc.get("siid", "?")
        name = svc.get("name", "?")
        desc = svc.get("description", "")
        stype = svc.get("type", "")
        print(f"{siid:>4} | {name:<22} | {desc:<20} | {stype}")


def main():
    parser = argparse.ArgumentParser(description="MIoT 批量创建工具（属性/方法/事件）")
    parser.add_argument("--excel", default="MIoT_属性创建模板.xlsx",
                        help="Excel 文件路径")
    parser.add_argument("--dry-run", action="store_true",
                        help="只解析不创建")
    parser.add_argument("--yes", "-y", action="store_true",
                        help="跳过确认直接创建")
    parser.add_argument("--skip-verify", action="store_true",
                        help="跳过验证步骤（推荐）")
    parser.add_argument("--delay", type=float, default=0.5,
                        help="每条创建之间的间隔秒数（默认0.5）")
    parser.add_argument("--list-services", action="store_true",
                        help="只列出产品服务，不创建")
    parser.add_argument("--siid", type=int, default=None,
                        help="只创建指定 siid 下的定义")
    parser.add_argument("--only", choices=["property", "action", "event"],
                        help="只创建指定类型（property/action/event）")
    args = parser.parse_args()

    # 读取 Excel
    if not os.path.exists(args.excel):
        print(f"❌ 文件不存在: {args.excel}")
        sys.exit(1)

    wb = load_workbook(args.excel)
    config = read_config(wb["公共配置"])

    # 读取各 Sheet（存在则读取，不存在则为空列表）
    props = read_properties(wb["属性定义"]) if "属性定义" in wb.sheetnames else []
    actions = read_actions(wb["方法定义"]) if "方法定义" in wb.sheetnames else []
    events = read_events(wb["事件定义"]) if "事件定义" in wb.sheetnames else []

    # 检查必填配置
    missing = []
    for key in ["serviceToken", "xiaomiiot_ph", "userId", "pdId", "model"]:
        if not config.get(key):
            missing.append(key)
    if missing:
        print(f"❌ 公共配置缺少必填项: {', '.join(missing)}")
        print("   请在 Excel「公共配置」Sheet 中填写")
        sys.exit(1)

    total_items = len(props) + len(actions) + len(events)
    if total_items == 0:
        print("❌ 没有找到任何属性/方法/事件定义")
        sys.exit(1)

    # 根据 --only 过滤
    if args.only == "property":
        actions, events = [], []
    elif args.only == "action":
        props, events = [], []
    elif args.only == "event":
        props, actions = [], []

    # 查询服务列表
    print("🔍 查询产品服务列表...")
    services = query_services(config)
    if not services:
        print("❌ 未查到服务列表，请检查 Cookie 和产品信息")
        sys.exit(1)

    print_services(services)

    if args.list_services:
        return

    # ── 解析属性任务 ──
    prop_tasks = []
    if props:
        print(f"\n📝 解析 {len(props)} 条属性定义...")
        for i, prop in enumerate(props):
            name = prop.get("name", f"行{i+1}")
            fmt = prop.get("format", "?")
            vtype = detect_value_type(fmt, prop)

            svc = match_service(prop, services)
            if svc:
                siid = svc["siid"]
                sname = svc.get("description", svc.get("name", ""))
            else:
                siid = prop.get("siid", "?")
                sname = "未匹配"

            if args.siid is not None and str(siid) != str(args.siid):
                continue

            body = build_request_body(prop, config, svc)
            prop_tasks.append({
                "index": i + 1, "name": name, "desc": prop.get("description", ""),
                "format": fmt, "value_type": vtype, "siid": siid,
                "service_name": sname, "body": body,
            })

        if prop_tasks:
            print(f"\n{'#':>3} | {'name':<20} | {'desc':<12} | {'format':<8} | {'vtype':<10} | siid | 服务")
            print("-" * 90)
            for t in prop_tasks:
                print(f"{t['index']:>3} | {t['name']:<20} | {t['desc']:<12} | "
                      f"{t['format']:<8} | {t['value_type']:<10} | {t['siid']:<4} | {t['service_name']}")

    # ── 解析方法任务 ──
    action_tasks = []
    if actions:
        print(f"\n📝 解析 {len(actions)} 条方法定义...")
        for i, item in enumerate(actions):
            name = item.get("name", f"行{i+1}")

            svc = match_service(item, services)
            if svc:
                siid = svc["siid"]
                sname = svc.get("description", svc.get("name", ""))
            else:
                siid = item.get("siid", "?")
                sname = "未匹配"

            if args.siid is not None and str(siid) != str(args.siid):
                continue

            body = build_action_request_body(item, config, svc)
            action_tasks.append({
                "index": i + 1, "name": name, "desc": item.get("description", ""),
                "siid": siid, "service_name": sname, "body": body,
            })

        if action_tasks:
            print(f"\n{'#':>3} | {'name':<20} | {'desc':<20} | siid | 服务")
            print("-" * 70)
            for t in action_tasks:
                print(f"{t['index']:>3} | {t['name']:<20} | {t['desc']:<20} | "
                      f"{t['siid']:<4} | {t['service_name']}")

    # ── 解析事件任务 ──
    event_tasks = []
    if events:
        print(f"\n📝 解析 {len(events)} 条事件定义...")
        for i, item in enumerate(events):
            name = item.get("name", f"行{i+1}")

            svc = match_service(item, services)
            if svc:
                siid = svc["siid"]
                sname = svc.get("description", svc.get("name", ""))
            else:
                siid = item.get("siid", "?")
                sname = "未匹配"

            if args.siid is not None and str(siid) != str(args.siid):
                continue

            body = build_event_request_body(item, config, svc)
            event_tasks.append({
                "index": i + 1, "name": name, "desc": item.get("description", ""),
                "siid": siid, "service_name": sname, "body": body,
            })

        if event_tasks:
            print(f"\n{'#':>3} | {'name':<20} | {'desc':<20} | siid | 服务")
            print("-" * 70)
            for t in event_tasks:
                print(f"{t['index']:>3} | {t['name']:<20} | {t['desc']:<20} | "
                      f"{t['siid']:<4} | {t['service_name']}")

    # 检查是否有任务
    total_tasks = len(prop_tasks) + len(action_tasks) + len(event_tasks)
    if total_tasks == 0:
        print("❌ 没有匹配的任务")
        sys.exit(1)

    # 汇总
    print(f"\n📋 任务汇总:")
    print(f"   属性: {len(prop_tasks)} 条")
    print(f"   方法: {len(action_tasks)} 条")
    print(f"   事件: {len(event_tasks)} 条")
    print(f"   合计: {total_tasks} 条")

    # 干跑模式
    if args.dry_run:
        print(f"\n🏁 干跑模式，共 {total_tasks} 条待创建（未实际执行）")
        for label, tasks, body_sample in [
            ("属性", prop_tasks, prop_tasks[0] if prop_tasks else None),
            ("方法", action_tasks, action_tasks[0] if action_tasks else None),
            ("事件", event_tasks, event_tasks[0] if event_tasks else None),
        ]:
            if body_sample:
                print(f"\n📄 {label}第1条请求体示例:")
                print(json.dumps(body_sample["body"], ensure_ascii=False, indent=2))
        return

    # 确认
    if not args.yes:
        ans = input(f"\n⚠️  即将创建 {total_tasks} 条（属性{len(prop_tasks)}+方法{len(action_tasks)}+事件{len(event_tasks)}），是否继续？[y/N] ").strip().lower()
        if ans not in ("y", "yes"):
            print("已取消")
            return

    # 批量创建
    all_results = []

    if prop_tasks:
        _, _, res = batch_create(prop_tasks, create_property, config, "属性", "piid", args.delay)
        all_results.extend(res)

    if action_tasks:
        _, _, res = batch_create(action_tasks, create_action, config, "方法", "aiid", args.delay)
        all_results.extend(res)

    if event_tasks:
        _, _, res = batch_create(event_tasks, create_event, config, "事件", "eiid", args.delay)
        all_results.extend(res)

    # 总汇总
    success = sum(1 for r in all_results if r["status"] == "success")
    failed = sum(1 for r in all_results if r["status"] != "success")
    print(f"\n{'='*50}")
    print(f"📊 全部完成: 成功 {success} / 失败 {failed} / 共 {total_tasks}")

    # 保存结果
    result_file = "miot_create_result.json"
    with open(result_file, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    print(f"\n💾 结果已保存到 {result_file}")


if __name__ == "__main__":
    main()
