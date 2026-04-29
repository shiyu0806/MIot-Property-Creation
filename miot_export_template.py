#!/usr/bin/env python3
"""
MIoT 模板导出工具
- 根据已有产品自动查询所有服务、属性、方法和事件
- 将数据自动填入 Excel 模板，直接可用于其他产品复用
- 用法: python3 miot_export_template.py --pid 33257 --model uwize.switch.yzw07 \
         --token <serviceToken> --ph <xiaomiiot_ph> --userid <userId>
"""

__all__ = [
    "query_services", "parse_prop_row", "parse_action_row", "parse_event_row",
    "write_prop_sheet", "write_action_sheet", "write_event_sheet",
    "write_config_sheet", "write_source_sheet",
]

import argparse
import json
import sys
import time

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from miot_common import (
    BASE as _BASE,
    DEFAULT_HEADERS as _DEFAULT_HEADERS,
    build_cookies as _build_cookies,
    build_params as _build_params,
    safe_request as _safe_request,
    safe_int as _safe_int,
    PROPERTY_COLUMNS as COLUMNS,
    ACTION_COLUMNS,
    EVENT_COLUMNS,
)

# ─── API ──────────────────────────────────────────────────────
BASE = _BASE  # 向后兼容
QUERY_PROPS_API    = f"{BASE}/cgi-std/api/v1/functionDefine/getInstanceProperties"
QUERY_ACTIONS_API  = f"{BASE}/cgi-std/api/v1/functionDefine/getInstanceActions"
QUERY_EVENTS_API   = f"{BASE}/cgi-std/api/v1/functionDefine/getInstanceEvents"
QUERY_SERVICES_API = f"{BASE}/cgi-std/api/v1/functionDefine/getInstanceServices"

HEADERS = dict(_DEFAULT_HEADERS)


# ─── API 请求 ─────────────────────────────────────────────────

def build_cookies(args) -> dict:
    return {
        "serviceToken": args.token,
        "userId":       args.userid,
        "xiaomiiot_ph": args.ph,
    }


def build_common_params(args) -> dict:
    return {
        "userId":      args.userid,
        "xiaomiiot_ph": args.ph,
        "pdId":        str(args.pid),
    }


def query_services(args) -> list[dict]:
    params = build_common_params(args)
    params.update({
        "model":       args.model,
        "connectType": str(args.connect_type),
        "language":    "zh_cn",
        "version":     "1",
        "status":      "0",
    })
    print(f"🔍 查询服务列表 (pdId={args.pid}, model={args.model})...")
    resp = requests.get(QUERY_SERVICES_API, params=params,
                        headers=HEADERS, cookies=build_cookies(args), timeout=15)
    data = resp.json()
    if data.get("status") != 200:
        print(f"❌ 查询服务失败: {json.dumps(data, ensure_ascii=False)}")
        sys.exit(1)
    services = data.get("result", [])
    print(f"✅ 共 {len(services)} 个服务")
    return services


def query_properties(siid: int, service_type: str, args) -> list[dict]:
    params = build_common_params(args)
    params.update({
        "version":     "1",
        "status":      "0",
        "siid":        str(siid),
        "serviceType": service_type,
        "model":       args.model,
        "connectType": str(args.connect_type),
        "language":    "zh_cn",
    })
    resp = requests.get(QUERY_PROPS_API, params=params,
                        headers=HEADERS, cookies=build_cookies(args), timeout=15)
    data = resp.json()
    if data.get("status") != 200:
        return []
    return data.get("result", [])


def query_actions(siid: int, service_type: str, args) -> list[dict]:
    """查询指定服务下的方法列表"""
    params = build_common_params(args)
    params.update({
        "version":     "1",
        "status":      "0",
        "siid":        str(siid),
        "serviceType": service_type,
        "model":       args.model,
        "connectType": str(args.connect_type),
        "language":    "zh_cn",
    })
    resp = requests.get(QUERY_ACTIONS_API, params=params,
                        headers=HEADERS, cookies=build_cookies(args), timeout=15)
    data = resp.json()
    if data.get("status") != 200:
        return []
    return data.get("result", [])


def query_events(siid: int, service_type: str, args) -> list[dict]:
    """查询指定服务下的事件列表"""
    params = build_common_params(args)
    params.update({
        "version":     "1",
        "status":      "0",
        "siid":        str(siid),
        "serviceType": service_type,
        "model":       args.model,
        "connectType": str(args.connect_type),
        "language":    "zh_cn",
    })
    resp = requests.get(QUERY_EVENTS_API, params=params,
                        headers=HEADERS, cookies=build_cookies(args), timeout=15)
    data = resp.json()
    if data.get("status") != 200:
        return []
    return data.get("result", [])


# ─── 属性数据解析 ─────────────────────────────────────────────

def detect_value_type(prop: dict) -> str:
    """
    推断属性的值类型:
      bool_range → format=bool
      enum       → valueList 非空
      number     → valueRange 非空
      string     → format=string
    """
    fmt = str(prop.get("format", "")).lower()
    if fmt == "bool":
        return "bool_range"
    if fmt == "string":
        return "string"
    if prop.get("valueList"):
        return "enum"
    return "number"


def format_value_list(value_list: list) -> str:
    """[{value:0,description:'关闭'}] → '0:关闭,1:开启'"""
    if not value_list:
        return ""
    parts = []
    for item in value_list:
        v = item.get("value", "")
        d = item.get("description", "")
        parts.append(f"{v}:{d}")
    return ",".join(parts)


def parse_prop_row(prop: dict, service: dict) -> dict:
    """
    将一条属性 API 数据解析为 Excel 行数据
    返回字段与 Excel 列定义对应
    """
    vtype = detect_value_type(prop)
    fmt = str(prop.get("format", "bool"))

    # valueList / valueRange
    if vtype == "enum":
        value_list_str  = format_value_list(prop.get("valueList", []))
        vr_min = vr_max = vr_step = ""
    elif vtype == "number":
        value_list_str = ""
        vr = prop.get("valueRange", [])
        vr_min  = vr[0] if len(vr) > 0 else 0
        vr_max  = vr[1] if len(vr) > 1 else 65535
        vr_step = vr[2] if len(vr) > 2 else 1
    else:
        value_list_str = ""
        vr_min = vr_max = vr_step = ""

    # access
    access_list = prop.get("access", ["read", "write", "notify"])
    access_str = ",".join(access_list) if isinstance(access_list, list) else str(access_list)

    return {
        "name":             prop.get("name", ""),
        "description":      prop.get("description", ""),
        "format":           fmt,
        "service_desc":     service.get("description", ""),
        "value_list":       value_list_str,
        "value_range_min":  vr_min,
        "value_range_max":  vr_max,
        "value_range_step": vr_step,
        "service_name":     service.get("name", ""),
        "siid":             service.get("siid", ""),
        "access":           access_str,
        "piid":             prop.get("piid", ""),
        # 原始元数据（用于参考，不写入属性定义 Sheet 正文，写到附录 Sheet）
        "_piid":            prop.get("piid", ""),
        "_siid":            service.get("siid", ""),
        "_service_type":    service.get("type", ""),
        "_value_type":      vtype,
    }


def parse_action_row(action: dict, service: dict) -> dict:
    """将一条方法 API 数据解析为 Excel 行数据"""
    return {
        "name":             action.get("name", ""),
        "description":      action.get("description", ""),
        "normalizationDesc": action.get("normalizationDesc", action.get("name", "")),
        "service_desc":     service.get("description", ""),
        "service_name":     service.get("name", ""),
        "siid":             service.get("siid", ""),
        "aiid":             action.get("aiid", ""),
        # 原始元数据
        "_aiid":            action.get("aiid", ""),
        "_siid":            service.get("siid", ""),
        "_service_type":    service.get("type", ""),
    }


def parse_event_row(event: dict, service: dict) -> dict:
    """将一条事件 API 数据解析为 Excel 行数据"""
    return {
        "name":             event.get("name", ""),
        "description":      event.get("description", ""),
        "normalizationDesc": event.get("normalizationDesc", event.get("name", "")),
        "service_desc":     service.get("description", ""),
        "service_name":     service.get("name", ""),
        "siid":             service.get("siid", ""),
        "eiid":             event.get("eiid", ""),
        # 原始元数据
        "_eiid":            event.get("eiid", ""),
        "_siid":            service.get("siid", ""),
        "_service_type":    service.get("type", ""),
    }


# ─── Excel 写入 ───────────────────────────────────────────────

# 样式
_header_font     = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_header_fill     = PatternFill("solid", fgColor="4472C4")
_opt_header_fill = PatternFill("solid", fgColor="8DB4E2")
_header_align    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_thin_border     = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_data_font       = Font(name="Arial", size=10)
_data_align      = Alignment(vertical="center", wrap_text=True)
_desc_font       = Font(name="Arial", size=9, color="666666")
_desc_fill       = PatternFill("solid", fgColor="D9E2F3")
_opt_desc_fill   = PatternFill("solid", fgColor="E8F0FE")

# COLUMNS, ACTION_COLUMNS, EVENT_COLUMNS 已从 miot_common 统一导入


def write_prop_sheet(ws, rows: list[dict]):
    """写入属性定义 Sheet（标题行 + 说明行 + 数据行）"""
    # 标题行
    for col_idx, (key, header, width, desc, required) in enumerate(COLUMNS, 1):
        fill = _header_fill if required else _opt_header_fill
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _header_font
        cell.fill = fill
        cell.alignment = _header_align
        cell.border = _thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 说明行
    for col_idx, (key, header, width, desc, required) in enumerate(COLUMNS, 1):
        d_fill = _desc_fill if required else _opt_desc_fill
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font = _desc_font
        cell.fill = d_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border

    # 数据行
    for row_idx, row_data in enumerate(rows, 3):
        for col_idx, (key, *_) in enumerate(COLUMNS, 1):
            val = row_data.get(key, "")
            # 空字符串的 int/float 值不写入
            if isinstance(val, str) and val == "":
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _data_font
            cell.alignment = _data_align
            cell.border = _thin_border

    # 数据验证 - format 列 (C列)
    format_dv = DataValidation(
        type="list",
        formula1='"bool,uint8,uint16,uint32,int8,int16,int32,float,string"',
        allow_blank=False,
    )
    ws.add_data_validation(format_dv)
    format_dv.add(f"C3:C{max(3, 2 + len(rows))}")

    # 冻结前两行
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A1:K{2 + len(rows)}"

    # 行高
    ws.row_dimensions[2].height = 55


def _write_generic_sheet(ws, rows: list[dict], columns: list[tuple],
                         header_fill_color: str = "4472C4",
                         opt_header_color: str = "8DB4E2",
                         desc_fill_color: str = "D9E2F3",
                         opt_desc_color: str = "E8F0FE"):
    """通用 Sheet 写入（标题行 + 说明行 + 数据行），供方法/事件复用"""
    h_fill = PatternFill("solid", fgColor=header_fill_color)
    oh_fill = PatternFill("solid", fgColor=opt_header_color)
    d_fill = PatternFill("solid", fgColor=desc_fill_color)
    od_fill = PatternFill("solid", fgColor=opt_desc_color)

    # 标题行
    for col_idx, (key, header, width, desc, required) in enumerate(columns, 1):
        fill = h_fill if required else oh_fill
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _header_font
        cell.fill = fill
        cell.alignment = _header_align
        cell.border = _thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 说明行
    for col_idx, (key, header, width, desc, required) in enumerate(columns, 1):
        df = d_fill if required else od_fill
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font = _desc_font
        cell.fill = df
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border

    # 数据行
    for row_idx, row_data in enumerate(rows, 3):
        for col_idx, (key, *_) in enumerate(columns, 1):
            val = row_data.get(key, "")
            if isinstance(val, str) and val == "":
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _data_font
            cell.alignment = _data_align
            cell.border = _thin_border

    # 冻结前两行
    ws.freeze_panes = "A3"
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{2 + len(rows)}"

    # 行高
    ws.row_dimensions[2].height = 40


def write_action_sheet(ws, rows: list[dict]):
    """写入方法定义 Sheet"""
    _write_generic_sheet(ws, rows, ACTION_COLUMNS,
                         header_fill_color="C55A11",
                         opt_header_color="F4B183",
                         desc_fill_color="FCE4D6",
                         opt_desc_color="FFF2CC")


def write_event_sheet(ws, rows: list[dict]):
    """写入事件定义 Sheet"""
    _write_generic_sheet(ws, rows, EVENT_COLUMNS,
                         header_fill_color="7030A0",
                         opt_header_color="B4A7D6",
                         desc_fill_color="E8D5F5",
                         opt_desc_color="F3E8FD")


def write_config_sheet(ws2, args):
    """写入公共配置 Sheet（已填入当前产品信息）"""
    config_headers = ["配置项", "值", "说明"]
    header_fill = PatternFill("solid", fgColor="548235")
    for col_idx, h in enumerate(config_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = _header_font
        cell.fill = header_fill
        cell.alignment = _header_align
        cell.border = _thin_border

    configs = [
        # 必填（已自动填入）
        ["userId",       args.userid,  "⚠️ 小米账号用户ID（必填）"],
        ["pdId",         "",           "⚠️ 目标产品ID（必填，请修改为目标产品）"],
        ["model",        "",           "⚠️ 目标设备型号（必填，请修改为目标产品）"],
        ["serviceToken", args.token,   "⚠️ 登录后从浏览器Cookie获取（必填）"],
        ["xiaomiiot_ph", args.ph,      "⚠️ 登录后从浏览器Cookie获取（必填）"],
        # 默认值
        ["connectType",  16,           "连接类型（默认16）"],
        ["language",     "zh_cn",      "语言（默认zh_cn）"],
        ["version",      1,            "版本号（默认1）"],
        ["status",       0,            "状态（默认0）"],
        ["source",       4,            "来源（默认4）"],
        ["standard",     "false",      "是否标准属性（默认false）"],
        # gattAccess 已自动等同于 access，无需单独配置
        ["access",       "read,write,notify", "默认访问权限，属性定义中可单独覆盖\n（gattAccess自动等同于access，无需单独配置）"],
    ]

    for row_idx, (key, val, desc) in enumerate(configs, 2):
        ws2.cell(row=row_idx, column=1, value=key).font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=row_idx, column=2, value=val).font = _data_font
        ws2.cell(row=row_idx, column=3, value=desc).font = Font(name="Arial", size=9, color="666666")
        for c in range(1, 4):
            ws2.cell(row=row_idx, column=c).border = _thin_border
            ws2.cell(row=row_idx, column=c).alignment = Alignment(vertical="center")

    # 标红必填项
    for row in range(2, 7):
        ws2.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=10, color="CC0000")
        ws2.cell(row=row, column=3).font = Font(name="Arial", size=9, color="CC0000")

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 50

    # 提示说明
    note_row = len(configs) + 3
    note_cell = ws2.cell(row=note_row, column=1,
                         value="💡 使用说明：pdId 和 model 请修改为目标产品信息，Cookie 如过期请重新获取")
    note_cell.font = Font(name="Arial", size=9, color="0070C0", italic=True)
    ws2.merge_cells(f"A{note_row}:C{note_row}")


def write_source_sheet(ws3, services: list[dict], all_rows: list[dict],
                       action_rows: list[dict] = None, event_rows: list[dict] = None):
    """写入原始数据参考 Sheet（服务列表 + 完整属性/方法/事件元数据）"""
    # 服务列表
    ws3.cell(row=1, column=1, value="=== 来源产品服务列表 ===").font = Font(bold=True, color="4472C4", size=11)
    svc_headers = ["siid", "name", "description", "type"]
    for col_idx, h in enumerate(svc_headers, 1):
        cell = ws3.cell(row=2, column=col_idx, value=h)
        cell.font = _header_font
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = _header_align
        cell.border = _thin_border

    for row_idx, svc in enumerate(services, 3):
        for col_idx, key in enumerate(["siid", "name", "description", "type"], 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=svc.get(key, ""))
            cell.font = _data_font
            cell.border = _thin_border

    offset = len(services) + 4

    # 属性元数据
    ws3.cell(row=offset, column=1, value="=== 完整属性元数据（原始） ===").font = Font(bold=True, color="4472C4", size=11)

    prop_headers = ["siid", "piid", "service_name", "service_desc", "service_type",
                    "name", "description", "format", "value_type",
                    "value_list", "value_range", "access"]
    for col_idx, h in enumerate(prop_headers, 1):
        cell = ws3.cell(row=offset + 1, column=col_idx, value=h)
        cell.font = _header_font
        cell.fill = PatternFill("solid", fgColor="7030A0")
        cell.alignment = _header_align
        cell.border = _thin_border

    for row_idx, r in enumerate(all_rows, offset + 2):
        vals = [
            r.get("_siid", ""),
            r.get("_piid", ""),
            r.get("service_name", ""),
            r.get("service_desc", ""),
            r.get("_service_type", ""),
            r.get("name", ""),
            r.get("description", ""),
            r.get("format", ""),
            r.get("_value_type", ""),
            r.get("value_list", ""),
            f"min={r.get('value_range_min','')} max={r.get('value_range_max','')} step={r.get('value_range_step','')}",
            r.get("access", ""),
        ]
        for col_idx, v in enumerate(vals, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=v)
            cell.font = _data_font
            cell.border = _thin_border

    # 方法元数据
    if action_rows:
        offset = offset + 2 + len(all_rows) + 2
        ws3.cell(row=offset, column=1, value="=== 完整方法元数据（原始） ===").font = Font(bold=True, color="C55A11", size=11)
        action_headers = ["siid", "aiid", "service_name", "service_desc", "service_type", "name", "description"]
        for col_idx, h in enumerate(action_headers, 1):
            cell = ws3.cell(row=offset + 1, column=col_idx, value=h)
            cell.font = _header_font
            cell.fill = PatternFill("solid", fgColor="C55A11")
            cell.alignment = _header_align
            cell.border = _thin_border
        for row_idx, r in enumerate(action_rows, offset + 2):
            vals = [
                r.get("_siid", ""),
                r.get("_aiid", ""),
                r.get("service_name", ""),
                r.get("service_desc", ""),
                r.get("_service_type", ""),
                r.get("name", ""),
                r.get("description", ""),
            ]
            for col_idx, v in enumerate(vals, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=v)
                cell.font = _data_font
                cell.border = _thin_border

    # 事件元数据
    if event_rows:
        offset = offset + 2 + (len(action_rows) if action_rows else 0) + 2
        ws3.cell(row=offset, column=1, value="=== 完整事件元数据（原始） ===").font = Font(bold=True, color="7030A0", size=11)
        event_headers = ["siid", "eiid", "service_name", "service_desc", "service_type", "name", "description"]
        for col_idx, h in enumerate(event_headers, 1):
            cell = ws3.cell(row=offset + 1, column=col_idx, value=h)
            cell.font = _header_font
            cell.fill = PatternFill("solid", fgColor="7030A0")
            cell.alignment = _header_align
            cell.border = _thin_border
        for row_idx, r in enumerate(event_rows, offset + 2):
            vals = [
                r.get("_siid", ""),
                r.get("_eiid", ""),
                r.get("service_name", ""),
                r.get("service_desc", ""),
                r.get("_service_type", ""),
                r.get("name", ""),
                r.get("description", ""),
            ]
            for col_idx, v in enumerate(vals, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=v)
                cell.font = _data_font
                cell.border = _thin_border

    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 6
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 20
    ws3.column_dimensions["E"].width = 40
    ws3.column_dimensions["F"].width = 22
    ws3.column_dimensions["G"].width = 20
    ws3.column_dimensions["H"].width = 10
    ws3.column_dimensions["I"].width = 12
    ws3.column_dimensions["J"].width = 35
    ws3.column_dimensions["K"].width = 25
    ws3.column_dimensions["L"].width = 22


# ─── 主流程 ──────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="从已有 MIoT 产品导出属性/方法/事件，生成可复用 Excel 模板",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python3 miot_export_template.py \\
    --pid 33257 \\
    --model uwize.switch.yzw07 \\
    --token 'v1hQ...' \\
    --ph 'CPAe...' \\
    --userid 1097752639
        """,
    )
    parser.add_argument("--pid",          required=True,  help="来源产品 ID（pdId）")
    parser.add_argument("--model",        required=True,  help="来源产品 model")
    parser.add_argument("--token",        required=True,  help="serviceToken（Cookie）")
    parser.add_argument("--ph",           required=True,  help="xiaomiiot_ph（Cookie）")
    parser.add_argument("--userid",       required=True,  help="userId")
    parser.add_argument("--connect-type", default=16,     type=int, help="connectType（默认16）")
    parser.add_argument("--output", "-o", default=None,   help="输出文件路径（默认自动生成）")
    parser.add_argument("--delay",        default=0.3,    type=float, help="服务间请求间隔（秒）")
    parser.add_argument("--json",         action="store_true", help="同时输出原始 JSON 数据")
    args = parser.parse_args()

    # 自动输出路径
    if not args.output:
        args.output = f"MIoT_模板_{args.model.replace('.', '_')}.xlsx"

    # 1. 查询服务列表
    services = query_services(args)

    print(f"\n{'siid':>4} | {'name':<22} | {'description':<22} | type")
    print("-" * 80)
    for svc in services:
        print(f"{svc.get('siid','?'):>4} | {svc.get('name',''):<22} | "
              f"{svc.get('description',''):<22} | {svc.get('type','')}")

    # 2. 逐个查询属性、方法、事件
    print(f"\n🔍 查询各服务属性、方法、事件...")
    all_rows = []
    action_rows = []
    event_rows = []
    for svc in services:
        siid = svc.get("siid")
        stype = svc.get("type", "")
        sname = svc.get("name", "?")
        sdesc = svc.get("description", "")

        # 属性
        props = query_properties(siid, stype, args)
        print(f"  siid={siid:<3} {sname:<22} ({sdesc}) → {len(props)} 个属性")
        for p in props:
            row = parse_prop_row(p, svc)
            all_rows.append(row)

        # 方法
        actions = query_actions(siid, stype, args)
        if actions:
            print(f"  siid={siid:<3} {sname:<22} ({sdesc}) → {len(actions)} 个方法")
        for a in actions:
            row = parse_action_row(a, svc)
            action_rows.append(row)

        # 事件
        events = query_events(siid, stype, args)
        if events:
            print(f"  siid={siid:<3} {sname:<22} ({sdesc}) → {len(events)} 个事件")
        for e in events:
            row = parse_event_row(e, svc)
            event_rows.append(row)

        if args.delay:
            time.sleep(args.delay)

    total_props = len(all_rows)
    total_actions = len(action_rows)
    total_events = len(event_rows)
    print(f"\n✅ 共抓取 {total_props} 条属性, {total_actions} 个方法, {total_events} 个事件，覆盖 {len(services)} 个服务")

    if total_props == 0 and total_actions == 0 and total_events == 0:
        print("⚠️  没有抓到任何数据，请检查 Cookie 是否有效或产品是否有定义")
        sys.exit(1)

    # 3. 可选输出 JSON
    if args.json:
        json_path = args.output.replace(".xlsx", ".json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({"properties": all_rows, "actions": action_rows, "events": event_rows},
                      f, ensure_ascii=False, indent=2)
        print(f"💾 原始数据已保存: {json_path}")

    # 4. 写入 Excel
    print(f"\n📊 生成 Excel 模板: {args.output}")
    wb = Workbook()

    # Sheet 1: 属性定义
    ws1 = wb.active
    ws1.title = "属性定义"
    write_prop_sheet(ws1, all_rows)

    # Sheet 2: 方法定义
    ws2 = wb.create_sheet("方法定义")
    write_action_sheet(ws2, action_rows)

    # Sheet 3: 事件定义
    ws3 = wb.create_sheet("事件定义")
    write_event_sheet(ws3, event_rows)

    # Sheet 4: 公共配置
    ws4 = wb.create_sheet("公共配置")
    write_config_sheet(ws4, args)

    # Sheet 5: 来源数据参考
    ws5 = wb.create_sheet("原始数据参考")
    write_source_sheet(ws5, services, all_rows, action_rows, event_rows)

    wb.save(args.output)
    print(f"✅ 模板已保存: {args.output}")
    print(f"\n💡 下一步：")
    print(f"   1. 打开 {args.output}")
    print(f"   2. 在「公共配置」Sheet 中修改 pdId 和 model 为目标产品")
    print(f"   3. 运行: .venv/bin/python miot_create_properties.py --excel {args.output} --dry-run")


if __name__ == "__main__":
    main()
