#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ========== 样式 ==========
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="4472C4")
opt_header_fill = PatternFill("solid", fgColor="8DB4E2")  # 可选列浅蓝色
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
data_font = Font(name="Arial", size=10)
data_align = Alignment(vertical="center", wrap_text=True)
desc_font = Font(name="Arial", size=9, color="666666")
desc_fill = PatternFill("solid", fgColor="D9E2F3")
opt_desc_fill = PatternFill("solid", fgColor="E8F0FE")  # 可选列说明行浅色
required_font = Font(name="Arial", size=9, color="CC0000", bold=True)

# ========== Sheet 1: 属性定义 ==========
ws = wb.active
ws.title = "属性定义"

# 列定义: (列名, 宽度, 说明, 是否必填)
# 必填列在前，可选列在后，用分隔行区分
columns = [
    # ── 必填 ──
    ("name",              20, "属性英文名\n如: on, mode, delay-time",              True),
    ("description",       25, "属性中文描述\n如: 开关, 模式, 延时时间",               True),
    ("format",            12, "数据格式\nbool/uint8/uint16/uint32\n/int8/int16/int32/float/string", True),
    ("service_desc",      22, "服务中文描述（推荐）\n如: 开关一键、按键1点动毫秒数\n用于匹配服务，优先级最高", True),
    # ── 条件必填（枚举/数值型需填） ──
    ("value_list",        35, "枚举值（仅enum类型）\n格式: 0:关闭,1:开启,2:待机\n非枚举留空", False),
    ("value_range_min",   14, "数值最小值\n（仅number类型）",                         False),
    ("value_range_max",   14, "数值最大值\n（仅number类型）",                         False),
    ("value_range_step",  14, "数值步长\n（仅number类型）",                           False),
    # ── 可选 ──
    ("service_name",      20, "服务英文名\n如: switch, jog-delay-time\n与service_desc配合可精确区分同名服务", False),
    ("siid",               8, "服务ID（备选）\n直接指定siid，填了则忽略service匹配",    False),
    ("access",            20, "访问权限\n默认: read,write,notify",                    False),
]

for col_idx, (name, width, desc, required) in enumerate(columns, 1):
    fill = header_fill if required else opt_header_fill
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.font = header_font
    cell.fill = fill
    cell.alignment = header_align
    cell.border = thin_border

    # 说明行
    d_fill = desc_fill if required else opt_desc_fill
    desc_cell = ws.cell(row=2, column=col_idx, value=desc)
    desc_cell.font = desc_font
    desc_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    desc_cell.fill = d_fill
    desc_cell.border = thin_border

    # 列宽
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# 示例数据（更简洁，只填必要列）
examples = [
    # bool 类型 — 最简填写
    ["on", "开关", "bool", "开关一键"],
    # 数值类型 — 需填 value_range
    ["delay-time", "延时时间", "uint32", "按键1点动毫秒数", "", 0, 65535, 1],
    # 枚举类型 — 需填 value_list
    ["mode", "模式", "uint8", "按键1点动毫秒数", "0:关闭,1:开启,2:待机"],
    # 用 siid 直接指定服务（备选方式）
    ["fault", "故障", "string", "", "", "", "", "", "", 2, "read,notify"],
]

for row_idx, row_data in enumerate(examples, 3):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border

# 数据验证 - format 列 (C列)
format_dv = DataValidation(
    type="list",
    formula1='"bool,uint8,uint16,uint32,int8,int16,int32,float,string"',
    allow_blank=False,
)
format_dv.error = "请选择有效的数据格式"
format_dv.errorTitle = "格式错误"
ws.add_data_validation(format_dv)
format_dv.add("C3:C200")

# 冻结前两行（标题+说明）
ws.freeze_panes = "A3"
# 自动筛选
ws.auto_filter.ref = f"A1:K{2 + len(examples)}"

# ========== Sheet 2: 公共配置 ==========
ws2 = wb.create_sheet("公共配置")

config_headers = ["配置项", "值", "说明"]
for col_idx, h in enumerate(config_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="548235")
    cell.alignment = header_align
    cell.border = thin_border

configs = [
    # ── 必填 ──
    ["userId",        "", "⚠️ 小米账号用户ID（必填）"],
    ["pdId",          "", "⚠️ 产品ID（必填）"],
    ["model",         "", "⚠️ 设备型号，如 uwize.switch.aiswi（必填）"],
    ["serviceToken",  "", "⚠️ 登录后从浏览器Cookie获取（必填）"],
    ["xiaomiiot_ph",  "", "⚠️ 登录后从浏览器Cookie获取（必填）"],
    # ── 有默认值 ──
    ["connectType",   16,  "连接类型（默认16）"],
    ["language",      "zh_cn", "语言（默认zh_cn）"],
    ["version",       1,   "版本号（默认1）"],
    ["status",        0,   "状态（默认0）"],
    ["source",        4,   "来源（默认4）"],
    ["standard",      "false", "是否标准属性（默认false）"],
    ["gattAccess",    "read,write,notify", "BLE访问权限（默认read,write,notify）"],
    ["access",        "read,write,notify", "默认访问权限，属性定义中可单独覆盖"],
]

for row_idx, (key, val, desc) in enumerate(configs, 2):
    key_cell = ws2.cell(row=row_idx, column=1, value=key)
    val_cell = ws2.cell(row=row_idx, column=2, value=val)
    desc_cell = ws2.cell(row=row_idx, column=3, value=desc)

    key_cell.font = Font(name="Arial", bold=True, size=10)
    val_cell.font = data_font
    desc_cell.font = Font(name="Arial", size=9, color="666666")

    for c in range(1, 4):
        ws2.cell(row=row_idx, column=c).border = thin_border
        ws2.cell(row=row_idx, column=c).alignment = Alignment(vertical="center")

# 标红必填项
for row in range(2, 7):  # 前5行是必填
    ws2.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=10, color="CC0000")
    ws2.cell(row=row, column=3).font = Font(name="Arial", size=9, color="CC0000")

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 50

# ========== Sheet 3: 填写说明 ==========
ws3 = wb.create_sheet("填写说明")

instructions = [
    ["项目", "说明"],
    ["快速开始", "1. 在「公共配置」填写必填项（标红项）\n2. 在「属性定义」每行填一个属性\n3. 至少填：name、description、format、service_desc\n4. 运行：python3 miot_create_properties.py --dry-run"],
    ["服务匹配", "脚本自动匹配属性到服务，优先级：\n1. siid 直接匹配（填了siid列则优先使用）\n2. service_desc 精确匹配（推荐，填服务中文名）\n3. service_name + service_desc 组合匹配\n4. service_name 精确匹配\n\n💡 提示：service_desc 是最简单的匹配方式，如「开关一键」「按键1点动毫秒数」"],
    ["格式类型 format", "bool: 布尔开关（自动设valueList=[] valueRange=[]）\nuint8/uint16/uint32: 无符号整数\nint8/int16/int32: 有符号整数\nfloat: 浮点数\nstring: 字符串"],
    ["枚举值 value_list", "仅枚举类型填写\n格式: 数值:描述,数值:描述\n示例: 0:关闭,1:开启,2:待机"],
    ["数值范围 value_range", "仅数值类型填写，三个字段：\nmin: 最小值（默认0）\nmax: 最大值（默认65535）\nstep: 步长（默认1）"],
    ["访问权限 access", "可选填，默认 read,write,notify\n多选用逗号分隔"],
    ["公共配置", "model、pdId、connectType 等全局参数统一在「公共配置」Sheet填写\n属性定义中无需重复填写"],
    ["Cookie获取", "1. 登录 iot.mi.com\n2. F12 → Application → Cookies\n3. 复制 serviceToken 和 xiaomiiot_ph 的值"],
    ["脚本命令", "# 查看服务列表\npython3 miot_create_properties.py --list-services\n# 干跑检查\npython3 miot_create_properties.py --dry-run\n# 正式创建\npython3 miot_create_properties.py --skip-verify -y"],
]

for row_idx, (title, content) in enumerate(instructions, 1):
    ws3.cell(row=row_idx, column=1, value=title).font = Font(name="Arial", bold=True, size=10)
    ws3.cell(row=row_idx, column=2, value=content).font = Font(name="Arial", size=10)
    ws3.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    for c in range(1, 3):
        ws3.cell(row=row_idx, column=c).border = thin_border

ws3.column_dimensions["A"].width = 18
ws3.column_dimensions["B"].width = 80

# 调整行高
for ws_sheet in [ws, ws2, ws3]:
    for row in ws_sheet.iter_rows():
        ws_sheet.row_dimensions[row[0].row].height = max(
            ws_sheet.row_dimensions[row[0].row].height or 20, 20
        )
ws.row_dimensions[2].height = 55

output = "/Users/shiyu/WorkBuddy/20260418171520/MIoT_属性创建模板.xlsx"
wb.save(output)
print(f"模板已保存: {output}")
