#!/usr/bin/env python3
"""
MIoT 平台工具（整合版）
功能：
  服务层：创建服务 / 导出服务
  属性层：导出模板 / 创建属性 / 生成模板
"""

import sys
import os
import json
import time
import traceback

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTabWidget,
    QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QTextEdit, QCheckBox,
    QFileDialog, QSpinBox, QGroupBox, QMessageBox, QProgressBar,
    QComboBox, QStatusBar, QDialog, QMenu, QSizePolicy, QListView,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSize
from PyQt6.QtGui import QFont, QIcon

# ── 属性层核心
from miot_export_template import (
    query_services as export_query_services,
    parse_prop_row,
    parse_action_row,
    parse_event_row,
    write_prop_sheet,
    write_action_sheet,
    write_event_sheet,
    write_config_sheet,
    write_source_sheet,
)
from miot_create_properties import (
    query_services as create_query_services,
    match_service,
    build_request_body,
    build_action_request_body,
    build_event_request_body,
    create_property,
    create_action,
    create_event,
    detect_value_type,
    read_config,
    read_properties,
    read_actions,
    read_events,
    batch_create,
    HEADERS,
    BASE,
    CREATE_PROP_API as CREATE_API,
    QUERY_SERVICES_API,
)
# ── 服务层核心
from miot_service_core import (
    get_services,
    sync_services,
    read_service_config_excel,
    read_service_list_excel,
    parse_service_str,
    check_product_status,
    modify_iid,
)
# ── 自动化核心
from miot_automation_core import (
    get_automation_list,
    check_standard_automation,
    save_automation,
    sync_automations,
    read_automation_excel,
    write_automation_export_excel,
)

import requests

# ── 登录模块
from miot_auth import (
    get_current_user, get_all_users, save_user, switch_user,
    remove_user, logout_current, MiLoginBrowser,
    update_user_group, get_curr_enterprise, get_enterprise_list,
    set_curr_enterprise,
)


# ─── 样式 ─────────────────────────────────────────────────────

STYLESHEET = """
QMainWindow { background-color: #f5f6fa; }
QTabWidget::pane {
    border: 1px solid #dcdde1;
    border-radius: 6px;
    background: white;
    margin-top: -1px;
}
QTabBar::tab {
    background: #dcdde1;
    padding: 10px 22px;
    margin-right: 2px;
    border-top-left-radius: 6px;
    border-top-right-radius: 6px;
    font-size: 13px;
    font-weight: bold;
    color: #555;
}
QTabBar::tab:selected {
    background: white;
    color: #2c3e50;
    border-bottom: 2px solid #e67e22;
}
QGroupBox {
    font-weight: bold;
    font-size: 13px;
    border: 1px solid #dcdde1;
    border-radius: 6px;
    margin-top: 12px;
    padding-top: 18px;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 12px;
    padding: 0 6px;
    color: #2c3e50;
}
QPushButton {
    background-color: #3498db;
    color: white;
    border: none;
    padding: 8px 20px;
    border-radius: 4px;
    font-size: 13px;
    font-weight: bold;
}
QPushButton:hover  { background-color: #2980b9; }
QPushButton:pressed { background-color: #2471a3; }
QPushButton:disabled { background-color: #bdc3c7; }
QPushButton#dangerBtn  { background-color: #e74c3c; }
QPushButton#dangerBtn:hover { background-color: #c0392b; }
QPushButton#successBtn { background-color: #27ae60; }
QPushButton#successBtn:hover { background-color: #229954; }
QPushButton#warnBtn  { background-color: #e67e22; }
QPushButton#warnBtn:hover { background-color: #ca6f1e; }
QLineEdit, QSpinBox, QComboBox {
    padding: 6px 10px;
    border: 1px solid #dcdde1;
    border-radius: 4px;
    font-size: 13px;
    background: white;
}
QLineEdit:focus, QSpinBox:focus { border-color: #3498db; }
QTextEdit {
    border: 1px solid #dcdde1;
    border-radius: 4px;
    font-family: "Menlo", "Consolas", monospace;
    font-size: 12px;
    background: #2c3e50;
    color: #ecf0f1;
    padding: 8px;
}
QProgressBar {
    border: 1px solid #dcdde1;
    border-radius: 4px;
    text-align: center;
    height: 22px;
    background: white;
}
QProgressBar::chunk { background-color: #3498db; border-radius: 3px; }
QLabel#titleLabel   { font-size: 18px; font-weight: bold; color: #2c3e50; }
QLabel#subtitleLabel { font-size: 12px; color: #7f8c8d; }
/* 企业下拉 */
QComboBox#entCombo {
    background-color: transparent;
    color: #2980b9;
    border: 2px solid #dcdde1;
    border-radius: 18px;
    padding: 4px 30px 4px 10px;
    font-size: 13px;
    font-weight: bold;
    min-height: 28px;
    max-width: 280px;
}
QComboBox#entCombo:hover { border-color: #3498db; }
QComboBox#entCombo::drop-down {
    subcontrol-origin: padding;
    subcontrol-position: right center;
    border: none;
    width: 24px;
}
QComboBox#entCombo::down-arrow {
    image: none;
    border-left: 5px solid transparent;
    border-right: 5px solid transparent;
    border-top: 6px solid #2980b9;
    margin-right: 8px;
}
QComboBox#entCombo QAbstractItemView {
    font-size: 13px;
    border: 1px solid #dcdde1;
    border-radius: 6px;
    selection-background-color: #eaf2f8;
    selection-color: #2c3e50;
    padding: 4px;
}
/* 用户区域 */
QPushButton#userBtn {
    background-color: transparent;
    color: #2c3e50;
    border: 2px solid #dcdde1;
    border-radius: 18px;
    padding: 4px 14px 4px 10px;
    font-size: 13px;
    font-weight: bold;
    min-height: 28px;
}
QPushButton#userBtn:hover { border-color: #3498db; color: #3498db; }
QPushButton#userBtn[loggedIn="true"] {
    border-color: #27ae60; color: #27ae60;
}
QPushButton#userBtn[loggedIn="true"]:hover {
    border-color: #e74c3c; color: #e74c3c;
}
QPushButton#loginBtn {
    background-color: #ff6700;
    color: white;
    border: none;
    border-radius: 4px;
    padding: 6px 16px;
    font-size: 13px;
    font-weight: bold;
}
QPushButton#loginBtn:hover { background-color: #e55d00; }
QTabWidget#innerTabs::pane {
    border: 1px solid #e0e0e0;
    border-radius: 4px;
    background: #fafafa;
}
QTabBar#innerTabs::tab {
    background: #e8e8e8;
    padding: 7px 18px;
    font-size: 12px;
    font-weight: normal;
}
QTabBar#innerTabs::tab:selected {
    background: #fafafa;
    color: #e67e22;
    border-bottom: 2px solid #e67e22;
}
"""


# ─── Worker Threads ───────────────────────────────────────────

class ExportPropWorker(QThread):
    """导出属性模板"""
    progress    = pyqtSignal(str)
    finished_ok  = pyqtSignal(str)
    finished_err = pyqtSignal(str)

    def __init__(self, pid, model, token, ph, userid, connect_type,
                 output_path, save_json, delay):
        super().__init__()
        self.pid = pid; self.model = model; self.token = token
        self.ph = ph; self.userid = userid; self.connect_type = connect_type
        self.output_path = output_path; self.save_json = save_json
        self.delay = delay
        self._cancel_flag = False

    def cancel(self):
        self._cancel_flag = True

    def run(self):
        try:
            cfg = {
                "userId": str(self.userid), "pdId": str(self.pid),
                "model": self.model, "serviceToken": self.token,
                "xiaomiiot_ph": self.ph,
                "connectType": str(self.connect_type),
            }
            cookies = {"serviceToken": self.token, "userId": str(self.userid),
                       "xiaomiiot_ph": self.ph}
            params_base = {"userId": str(self.userid),
                           "xiaomiiot_ph": self.ph, "pdId": str(self.pid)}

            all_props = []
            all_actions = []
            all_events = []
            max_retries = 3

            def _safe_get(url, params, label, retries=max_retries):
                """带重试的 GET 请求，返回 result 列表"""
                for attempt in range(1, retries + 1):
                    try:
                        r = requests.get(url, params=params, headers=HEADERS, cookies=cookies, timeout=15)
                        data = r.json()
                        return data.get("result", []) if data.get("status") == 200 else []
                    except Exception as e:
                        if attempt < retries:
                            self.progress.emit(f"  ⚠️ {label}查询失败(第{attempt}次)，1s后重试: {e}")
                            time.sleep(1)
                        else:
                            self.progress.emit(f"  ❌ {label}查询失败(已重试{retries}次): {e}")
                            raise

            self.progress.emit("📋 正在查询产品服务列表...")
            params = {**params_base, "model": self.model,
                      "connectType": str(self.connect_type),
                      "language": "zh_cn", "version": "1", "status": "0"}
            services = None
            for attempt in range(1, max_retries + 1):
                try:
                    resp = requests.get(QUERY_SERVICES_API, params=params,
                                        headers=HEADERS, cookies=cookies, timeout=15)
                    if resp.status_code != 200:
                        self.finished_err.emit(f"HTTP 请求失败 (status={resp.status_code})\n请检查网络连接")
                        return
                    try:
                        data = resp.json()
                    except Exception:
                        snippet = resp.text[:500] if resp.text else "(空响应)"
                        self.finished_err.emit(f"API 返回非 JSON 内容 (HTTP {resp.status_code}):\n{snippet}\n\n常见原因：Cookie 过期，请重新获取 serviceToken 和 xiaomiiot_ph")
                        return
                    if data.get("status") != 200:
                        self.finished_err.emit(f"查询服务失败: {data}")
                        return
                    services = data.get("result", [])
                    break
                except Exception as e:
                    if attempt < max_retries:
                        self.progress.emit(f"  ⚠️ 服务列表查询失败(第{attempt}次)，1s后重试: {e}")
                        time.sleep(1)
                    else:
                        self.finished_err.emit(f"❌ 服务列表查询失败(已重试{max_retries}次): {e}")
                        return
            if not services:
                self.finished_err.emit("未查到服务，请检查 Cookie 和产品信息")
                return
            self.progress.emit(f"✅ 找到 {len(services)} 个服务")

            for i, svc in enumerate(services):
                if self._cancel_flag:
                    self.finished_err.emit("⚠️ 用户取消")
                    return
                siid = svc.get("siid", "?")
                sname = svc.get("description", svc.get("name", ""))
                stype = svc.get("type", "")
                self.progress.emit(f"🔍 [{i+1}/{len(services)}] 查询服务 siid={siid} ({sname})...")
                params2 = {**params_base, "version": "1", "status": "0",
                           "siid": str(siid), "serviceType": stype,
                           "model": self.model,
                           "connectType": str(self.connect_type), "language": "zh_cn"}

                # 属性
                props = _safe_get(
                    "https://iot.mi.com/cgi-std/api/v1/functionDefine/getInstanceProperties",
                    params2, "属性")
                for p in props:
                    p["_service"] = svc
                all_props.extend(props)

                # 方法
                actions = _safe_get(
                    "https://iot.mi.com/cgi-std/api/v1/functionDefine/getInstanceActions",
                    params2, "方法")
                for a in actions:
                    a["_service"] = svc
                all_actions.extend(actions)

                # 事件
                events = _safe_get(
                    "https://iot.mi.com/cgi-std/api/v1/functionDefine/getInstanceEvents",
                    params2, "事件")
                for e in events:
                    e["_service"] = svc
                all_events.extend(events)

                if self.delay > 0:
                    time.sleep(self.delay)

            self.progress.emit(f"✅ 共获取 {len(all_props)} 条属性, {len(all_actions)} 个方法, {len(all_events)} 个事件")

            if not self.output_path:
                safe_model = self.model.replace(".", "_").replace("-", "_")
                self.output_path = os.path.join(os.path.expanduser("~"), "Desktop", f"MIoT_模板_{safe_model}.xlsx")

            self.progress.emit("📝 正在生成 Excel 模板...")
            from openpyxl import Workbook
            wb = Workbook()
            ws1 = wb.active; ws1.title = "属性定义"
            rows_data = [parse_prop_row(p, p.get("_service", {})) for p in all_props]
            write_prop_sheet(ws1, rows_data)

            # 方法定义 Sheet
            action_rows = [parse_action_row(a, a.get("_service", {})) for a in all_actions]
            ws2 = wb.create_sheet("方法定义")
            write_action_sheet(ws2, action_rows)

            # 事件定义 Sheet
            event_rows = [parse_event_row(e, e.get("_service", {})) for e in all_events]
            ws3 = wb.create_sheet("事件定义")
            write_event_sheet(ws3, event_rows)

            # 公共配置 Sheet
            ws4 = wb.create_sheet("公共配置")

            class _Args:
                pass
            args = _Args()
            args.pid = self.pid; args.model = self.model
            args.token = self.token; args.ph = self.ph
            args.userid = self.userid; args.connect_type = self.connect_type
            write_config_sheet(ws4, args)

            ws5 = wb.create_sheet("原始数据参考")
            write_source_sheet(ws5, services, rows_data, action_rows, event_rows)
            wb.save(self.output_path)

            if self.save_json:
                json_path = self.output_path.replace(".xlsx", ".json")
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump({"services": services, "properties": all_props,
                               "actions": all_actions, "events": all_events},
                              f, ensure_ascii=False, indent=2, default=str)
                self.progress.emit(f"💾 JSON 已保存: {json_path}")

            self.finished_ok.emit(self.output_path)
        except Exception:
            self.finished_err.emit(f"导出失败:\n{traceback.format_exc()}")


class CreatePropWorker(QThread):
    """批量创建属性"""
    progress         = pyqtSignal(str)
    update_progress  = pyqtSignal(int, int)
    finished_ok      = pyqtSignal(int, int)
    finished_err     = pyqtSignal(str)

    def __init__(self, config, props, services, delay):
        super().__init__()
        self.config = config; self.props = props
        self.services = services; self.delay = delay
        self._cancel_flag = False

    def cancel(self):
        self._cancel_flag = True

    def run(self):
        try:
            success = failed = 0
            results = []
            for i, prop in enumerate(self.props):
                if self._cancel_flag:
                    self.progress.emit("⚠️ 用户取消，已停止创建")
                    break
                name = prop.get("name", f"行{i+1}")
                svc = match_service(prop, self.services)
                siid = svc["siid"] if svc else prop.get("siid", "?")
                sname = svc.get("description", svc.get("name", "")) if svc else "未匹配"
                body = build_request_body(prop, self.config, svc)
                self.update_progress.emit(i + 1, len(self.props))
                self.progress.emit(f"  [{i+1}] {name} → siid={siid} ({sname}) ...")
                try:
                    resp = create_property(body, self.config)
                    if resp.get("status") == 200:
                        piid = resp.get("result")
                        # 校验并修正 piid
                        expected_piid = prop.get("piid")
                        if expected_piid and str(expected_piid).strip():
                            try:
                                expected_piid_int = int(expected_piid)
                                if int(piid) != expected_piid_int:
                                    self.progress.emit(f"    🔧 PIID {piid}→{expected_piid_int} 修正中...")
                                    r = modify_iid(self.config, siid, piid, expected_piid_int, "PIID")
                                    if r.get("code") == 0 or r.get("status") == 200:
                                        self.progress.emit(f"  ✅ {name} 成功 (piid={expected_piid_int}, 已修正)")
                                        success += 1
                                        results.append({"name": name, "status": "success", "piid": expected_piid_int, "original_piid": piid, "siid": siid})
                                    else:
                                        msg_m = r.get("message", r.get("msg", str(r)))
                                        self.progress.emit(f"  ⚠️ {name} 修正失败: {msg_m}")
                                        self.progress.emit(f"  ✅ {name} 成功 (piid={piid})")
                                        success += 1
                                        results.append({"name": name, "status": "success", "piid": piid, "modify_error": msg_m, "siid": siid})
                                else:
                                    self.progress.emit(f"  ✅ {name} 成功 (piid={piid})")
                                    success += 1
                                    results.append({"name": name, "status": "success", "piid": piid, "siid": siid})
                            except (ValueError, TypeError):
                                self.progress.emit(f"  ✅ {name} 成功 (piid={piid})")
                                success += 1
                                results.append({"name": name, "status": "success", "piid": piid, "siid": siid})
                        else:
                            self.progress.emit(f"  ✅ {name} 成功 (piid={piid})")
                            success += 1
                            results.append({"name": name, "status": "success", "piid": piid, "siid": siid})
                    else:
                        msg = resp.get("message", resp.get("msg", json.dumps(resp, ensure_ascii=False)))
                        self.progress.emit(f"  ❌ {name} 失败 ({msg})")
                        failed += 1
                        results.append({"name": name, "status": "failed", "error": msg, "siid": siid})
                except Exception as e:
                    self.progress.emit(f"  ❌ {name} 异常 ({e})")
                    failed += 1
                    results.append({"name": name, "status": "error", "error": str(e), "siid": siid})
                if self.delay > 0:
                    time.sleep(self.delay)

            result_path = os.path.join(os.path.expanduser("~"), "Desktop", "miot_create_result.json")
            with open(result_path, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            self.finished_ok.emit(success, failed)
        except Exception:
            self.finished_err.emit(f"创建失败:\n{traceback.format_exc()}")


class CreateAllWorker(QThread):
    """批量创建属性+方法+事件"""
    progress         = pyqtSignal(str)
    update_progress  = pyqtSignal(int, int)
    finished_ok      = pyqtSignal(int, int)
    finished_err     = pyqtSignal(str)

    def __init__(self, config, tasks, services, delay):
        """
        tasks: [(type_label, item, build_fn, create_fn, id_field, svc), ...]
        """
        super().__init__()
        self.config = config; self.tasks = tasks
        self.services = services; self.delay = delay
        self._cancel_flag = False

    def cancel(self):
        self._cancel_flag = True

    def run(self):
        try:
            success = failed = 0
            results = []
            for i, (type_label, item, build_fn, create_fn, id_field, svc) in enumerate(self.tasks):
                if self._cancel_flag:
                    self.progress.emit("⚠️ 用户取消，已停止创建")
                    break
                name = item.get("name", f"行{i+1}")
                siid = svc["siid"] if svc else item.get("siid", "?")
                sname = svc.get("description", svc.get("name", "")) if svc else "未匹配"
                body = build_fn(item, self.config, svc)
                self.update_progress.emit(i + 1, len(self.tasks))
                self.progress.emit(f"  [{i+1}][{type_label}] {name} → siid={siid} ({sname}) ...")
                try:
                    resp = create_fn(body, self.config)
                    if resp.get("status") == 200:
                        new_id = resp.get("result")
                        # 校验并修正 ID
                        expected_id = item.get(id_field)
                        which_iid_map = {"piid": "PIID", "aiid": "AIID", "eiid": "EIID"}
                        which_iid = which_iid_map.get(id_field, "")
                        if expected_id and which_iid and str(expected_id).strip():
                            try:
                                expected_id_int = int(expected_id)
                                if int(new_id) != expected_id_int:
                                    self.progress.emit(f"    🔧 {which_iid} {new_id}→{expected_id_int} 修正中...")
                                    r = modify_iid(self.config, siid, new_id, expected_id_int, which_iid)
                                    if r.get("code") == 0 or r.get("status") == 200:
                                        self.progress.emit(f"  ✅ [{type_label}] {name} 成功 ({id_field}={expected_id_int}, 已修正)")
                                        success += 1
                                        results.append({"type": type_label, "name": name, "status": "success", id_field: expected_id_int, "original_id": new_id, "siid": siid})
                                    else:
                                        msg_m = r.get("message", r.get("msg", str(r)))
                                        self.progress.emit(f"  ⚠️ [{type_label}] {name} 修正失败: {msg_m}")
                                        self.progress.emit(f"  ✅ [{type_label}] {name} 成功 ({id_field}={new_id})")
                                        success += 1
                                        results.append({"type": type_label, "name": name, "status": "success", id_field: new_id, "modify_error": msg_m, "siid": siid})
                                else:
                                    self.progress.emit(f"  ✅ [{type_label}] {name} 成功 ({id_field}={new_id})")
                                    success += 1
                                    results.append({"type": type_label, "name": name, "status": "success", id_field: new_id, "siid": siid})
                            except (ValueError, TypeError):
                                self.progress.emit(f"  ✅ [{type_label}] {name} 成功 ({id_field}={new_id})")
                                success += 1
                                results.append({"type": type_label, "name": name, "status": "success", id_field: new_id, "siid": siid})
                        else:
                            self.progress.emit(f"  ✅ [{type_label}] {name} 成功 ({id_field}={new_id})")
                            success += 1
                            results.append({"type": type_label, "name": name, "status": "success", id_field: new_id, "siid": siid})
                    else:
                        msg = resp.get("message", resp.get("msg", json.dumps(resp, ensure_ascii=False)))
                        self.progress.emit(f"  ❌ [{type_label}] {name} 失败 ({msg})")
                        failed += 1
                        results.append({"type": type_label, "name": name, "status": "failed", "error": msg, "siid": siid})
                except Exception as e:
                    self.progress.emit(f"  ❌ [{type_label}] {name} 异常 ({e})")
                    failed += 1
                    results.append({"type": type_label, "name": name, "status": "error", "error": str(e), "siid": siid})
                if self.delay > 0:
                    time.sleep(self.delay)

            result_path = os.path.join(os.path.expanduser("~"), "Desktop", "miot_create_result.json")
            with open(result_path, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            self.finished_ok.emit(success, failed)
        except Exception:
            self.finished_err.emit(f"创建失败:\n{traceback.format_exc()}")


class SyncServiceWorker(QThread):
    """批量同步服务（创建 / 修正 siid）"""
    progress    = pyqtSignal(str)
    finished_ok  = pyqtSignal(dict)
    finished_err = pyqtSignal(str)

    def __init__(self, config, service_rows, dry_run, delay=0.5):
        super().__init__()
        self.config = config; self.service_rows = service_rows
        self.dry_run = dry_run; self._cancel = False
        self.delay = delay

    def cancel(self):
        self._cancel = True

    def run(self):
        try:
            result = sync_services(
                self.config, self.service_rows,
                dry_run=self.dry_run,
                log_fn=self.progress.emit,
                cancelled_fn=lambda: self._cancel,
            )
            results_path = os.path.join(os.path.expanduser("~"), "Desktop", "sync_results.json")
            with open(results_path, "w", encoding="utf-8") as f:
                json.dump(result["results"], f, ensure_ascii=False, indent=2)
            self.finished_ok.emit(result)
        except Exception:
            self.finished_err.emit(f"同步失败:\n{traceback.format_exc()}")


# ─── 自动化 Worker ────────────────────────────────────────────

class ExportAutomationWorker(QThread):
    """导出自动化列表"""
    progress    = pyqtSignal(str)
    finished_ok  = pyqtSignal(str)
    finished_err = pyqtSignal(str)

    def __init__(self, config, output_path):
        super().__init__()
        self.config = config; self.output_path = output_path

    def run(self):
        try:
            self.progress.emit("📋 正在查询自动化列表...")
            auto_list = get_automation_list(self.config)
            then_count = sum(1 for a in auto_list if a.get("_trType") == "then")
            if_count = sum(1 for a in auto_list if a.get("_trType") == "if")
            then_action = sum(1 for a in auto_list if a.get("_trType") == "then" and a.get("actionList"))
            then_simple = then_count - then_action
            self.progress.emit(f"✅ 找到 {len(auto_list)} 个自动化（执行动作: {then_count}[组合{then_action}+普通{then_simple}], 触发条件: {if_count}）")

            if not auto_list:
                self.finished_err.emit("未查到自动化，请检查 Cookie 和产品信息")
                return

            self.progress.emit("📝 正在生成 Excel...")
            write_automation_export_excel(self.output_path, self.config, auto_list)
            self.finished_ok.emit(self.output_path)
        except Exception:
            self.finished_err.emit(f"导出失败:\n{traceback.format_exc()}")


class CreateAutomationWorker(QThread):
    """批量创建自定义自动化"""
    progress    = pyqtSignal(str)
    update_progress = pyqtSignal(int, int)  # current, total
    finished_ok  = pyqtSignal(int, int)     # success, failed
    finished_err = pyqtSignal(str)

    def __init__(self, config, auto_items, dry_run=False, delay=0.5):
        super().__init__()
        self.config = config; self.auto_items = auto_items
        self.dry_run = dry_run; self.delay = delay
        self._cancel = False

    def cancel(self):
        self._cancel = True

    def run(self):
        try:
            result = sync_automations(
                self.config, self.auto_items,
                dry_run=self.dry_run,
                delay=self.delay,
                log_fn=self.progress.emit,
                cancelled_fn=lambda: self._cancel,
            )
            s = len(result["success"])
            f = len(result["failed"])
            # 保存结果
            results_path = os.path.join(os.path.expanduser("~"), "Desktop", "automation_results.json")
            with open(results_path, "w", encoding="utf-8") as fout:
                json.dump(result, fout, ensure_ascii=False, indent=2, default=str)
            self.finished_ok.emit(s, f)
        except Exception:
            self.finished_err.emit(f"创建失败:\n{traceback.format_exc()}")


class ExportServiceWorker(QThread):
    """导出服务 / 属性详情"""
    progress    = pyqtSignal(str)
    finished_ok  = pyqtSignal(str)
    finished_err = pyqtSignal(str)

    def __init__(self, config, output_path, export_props=False):
        super().__init__()
        self.config = config; self.output_path = output_path
        self.export_props = export_props

    def run(self):
        try:
            import pandas as pd
            self.progress.emit(f"正在获取 {self.config.get('model')} 的服务列表...")
            services = get_services(self.config)
            self.progress.emit(f"✅ 获取到 {len(services)} 个服务")

            config_rows = [
                {"参数名": "userId",       "值": self.config.get("userId", "")},
                {"参数名": "xiaomiiot_ph", "值": self.config.get("xiaomiiot_ph", "")},
                {"参数名": "serviceToken", "值": self.config.get("serviceToken", "")},
                {"参数名": "pdId",         "值": self.config.get("pdId", "")},
                {"参数名": "model",        "值": self.config.get("model", "")},
            ]
            df_config = pd.DataFrame(config_rows)

            svc_rows = []
            prop_rows = []
            for svc in services:
                siid  = svc.get("siid", "")
                sname = svc.get("name", "")
                sdesc = svc.get("description", "") or svc.get("normalizationDesc", "")
                ndesc = svc.get("normalizationDesc", "")
                std   = "true" if svc.get("standard") else "false"
                svc_rows.append({
                    "服务ID": siid, "服务名称": sname, "服务描述": sdesc,
                    "标准化描述": ndesc, "是否标准服务": std,
                })
                if self.export_props:
                    parsed = parse_service_str(svc)
                    for prop in parsed["properties"]:
                        ptype = prop.get("type", "")
                        pname = ptype.split(":")[-2] if ":" in ptype else ""
                        prop_rows.append({
                            "siid": siid, "服务名称": sname,
                            "piid": prop.get("iid", ""), "类型": "属性",
                            "属性名称": pname, "描述": prop.get("description", ""),
                            "格式": prop.get("format", ""),
                            "访问权限": ",".join(prop.get("access", [])),
                            "值列表": json.dumps(prop.get("value-list", []), ensure_ascii=False) if prop.get("value-list") else "",
                            "值范围": json.dumps(prop.get("value-range", []), ensure_ascii=False) if prop.get("value-range") else "",
                        })
                    for evt in parsed["events"]:
                        etype = evt.get("type", "")
                        prop_rows.append({
                            "siid": siid, "服务名称": sname,
                            "piid": evt.get("iid", ""), "类型": "事件",
                            "属性名称": etype.split(":")[-2] if ":" in etype else "",
                            "描述": evt.get("description", ""),
                            "格式": "", "访问权限": "", "值列表": "", "值范围": "",
                        })
                    for act in parsed["actions"]:
                        atype = act.get("type", "")
                        prop_rows.append({
                            "siid": siid, "服务名称": sname,
                            "piid": act.get("iid", ""), "类型": "动作",
                            "属性名称": atype.split(":")[-2] if ":" in atype else "",
                            "描述": act.get("description", ""),
                            "格式": "", "访问权限": "", "值列表": "", "值范围": "",
                        })
                    self.progress.emit(
                        f"  siid={siid} {sname}  属性:{len(parsed['properties'])} "
                        f"事件:{len(parsed['events'])} 动作:{len(parsed['actions'])}"
                    )

            df_svc = pd.DataFrame(svc_rows)
            with pd.ExcelWriter(self.output_path, engine="openpyxl") as writer:
                df_config.to_excel(writer, index=False, sheet_name="产品配置")
                df_svc.to_excel(writer, index=False, sheet_name="服务列表")
                if self.export_props and prop_rows:
                    pd.DataFrame(prop_rows).to_excel(writer, index=False, sheet_name="属性详情")

            self.finished_ok.emit(self.output_path)
        except Exception:
            self.finished_err.emit(f"导出失败:\n{traceback.format_exc()}")


# ─── 公共小组件 ───────────────────────────────────────────────

def _make_log_panel(parent_layout) -> QTextEdit:
    lbl = QLabel("运行日志")
    lbl.setStyleSheet("font-weight: bold; font-size: 13px;")
    log = QTextEdit()
    log.setReadOnly(True)
    parent_layout.addWidget(lbl)
    parent_layout.addWidget(log)
    return log

def _make_progress(parent_layout) -> QProgressBar:
    pb = QProgressBar()
    pb.setVisible(False)
    parent_layout.addWidget(pb)
    return pb

def _inject_group_id(config: dict):
    """从当前登录用户自动注入 groupId 到 config（如果 config 中没有的话）"""
    if config.get("groupId"):
        return
    cur = get_current_user()
    if cur and cur.get("groupId"):
        config["groupId"] = cur["groupId"]

def _cookie_group(parent_layout, prefix: str, show_userid=True):
    """
    返回 (grp, token_edit, ph_edit, userid_edit_or_None)
    prefix 用于内部区分，不展示给用户
    如果已登录，自动填充 Cookie 字段
    """
    grp = QGroupBox("Cookie 信息")
    form = QFormLayout()
    token = QLineEdit(); token.setEchoMode(QLineEdit.EchoMode.Password)
    token.setPlaceholderText("浏览器 Cookie 中的 serviceToken")
    ph = QLineEdit(); ph.setEchoMode(QLineEdit.EchoMode.Password)
    ph.setPlaceholderText("浏览器 Cookie 中的 xiaomiiot_ph")
    userid_edit = None
    form.addRow("serviceToken:", token)
    form.addRow("xiaomiiot_ph:", ph)
    if show_userid:
        userid_edit = QLineEdit()
        userid_edit.setPlaceholderText("如 1097752639")
        form.addRow("userId:", userid_edit)

    # 自动填充当前用户的 Cookie
    cur = get_current_user()
    if cur:
        token.setText(cur.get("serviceToken", ""))
        ph.setText(cur.get("xiaomiiot_ph", ""))
        if userid_edit:
            userid_edit.setText(cur.get("userId", ""))

    chk = QCheckBox("显示 Cookie")
    def toggle(checked):
        mode = QLineEdit.EchoMode.Normal if checked else QLineEdit.EchoMode.Password
        token.setEchoMode(mode); ph.setEchoMode(mode)
    chk.toggled.connect(toggle)
    form.addRow("", chk)
    grp.setLayout(form)
    parent_layout.addWidget(grp)
    return grp, token, ph, userid_edit


# ─── Tab: 创建服务 ────────────────────────────────────────────

class CreateServiceTab(QWidget):
    def __init__(self):
        super().__init__()
        self._worker = None
        self._build()

    def _build(self):
        layout = QHBoxLayout(self)

        # 左侧表单
        left = QWidget(); left.setFixedWidth(460)
        lv = QVBoxLayout(left)

        # Excel 文件
        grp_file = QGroupBox("Excel 文件（服务模板）")
        fv = QHBoxLayout()
        self.file_edit = QLineEdit(); self.file_edit.setPlaceholderText("选择服务 Excel 模板")
        btn_browse = QPushButton("浏览...")
        btn_browse.clicked.connect(self._browse_file)
        fv.addWidget(self.file_edit); fv.addWidget(btn_browse)
        grp_file.setLayout(fv)
        lv.addWidget(grp_file)

        # 产品信息覆盖
        grp_prod = QGroupBox("产品信息（可覆盖 Excel 配置）")
        form_prod = QFormLayout()
        self.pid_edit = QLineEdit(); self.pid_edit.setPlaceholderText("留空使用 Excel 配置")
        self.model_edit = QLineEdit(); self.model_edit.setPlaceholderText("留空使用 Excel 配置")
        form_prod.addRow("产品ID (pdId):", self.pid_edit)
        form_prod.addRow("产品型号 (model):", self.model_edit)
        grp_prod.setLayout(form_prod)
        lv.addWidget(grp_prod)

        # Cookie 覆盖
        _, self.token_edit, self.ph_edit, self.userid_edit = _cookie_group(lv, "svc_crt")
        self.token_edit.setPlaceholderText("留空使用 Excel 配置")
        self.ph_edit.setPlaceholderText("留空使用 Excel 配置")
        self.userid_edit.setPlaceholderText("留空使用 Excel 配置")

        # 选项
        grp_opt = QGroupBox("选项")
        form_opt = QFormLayout()
        self.delay_spin = QSpinBox()
        self.delay_spin.setRange(100, 2000); self.delay_spin.setValue(500)
        self.delay_spin.setSingleStep(100); self.delay_spin.setSuffix(" ms")
        form_opt.addRow("请求间隔:", self.delay_spin)
        grp_opt.setLayout(form_opt)
        lv.addWidget(grp_opt)

        # 按钮
        btn_row = QHBoxLayout()
        self.btn_dry = QPushButton("🧪 干跑检查")
        self.btn_dry.clicked.connect(self._start_dry)
        self.btn_run = QPushButton("🚀 开始创建")
        self.btn_run.setObjectName("warnBtn")
        self.btn_run.clicked.connect(self._start_create)
        self.btn_cancel = QPushButton("取消")
        self.btn_cancel.clicked.connect(self._cancel)
        self.btn_cancel.setEnabled(False)
        btn_row.addWidget(self.btn_dry); btn_row.addWidget(self.btn_run)
        btn_row.addWidget(self.btn_cancel)
        lv.addLayout(btn_row)
        lv.addStretch()

        # 右侧日志
        right = QWidget(); rv = QVBoxLayout(right)
        self.log = _make_log_panel(rv)
        self.progress = _make_progress(rv)

        layout.addWidget(left)
        layout.addWidget(right, stretch=1)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择服务 Excel", "", "Excel (*.xlsx *.xls)")
        if path:
            self.file_edit.setText(path)

    def _build_config(self):
        path = self.file_edit.text().strip()
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "提示", "请选择有效的 Excel 文件")
            return None, None
        try:
            config = read_service_config_excel(path)
            rows   = read_service_list_excel(path)
        except Exception as e:
            QMessageBox.critical(self, "读取失败", str(e))
            return None, None

        # 覆盖
        for key, widget in [
            ("pdId", self.pid_edit), ("model", self.model_edit),
            ("serviceToken", self.token_edit),
            ("xiaomiiot_ph", self.ph_edit), ("userId", self.userid_edit),
        ]:
            val = widget.text().strip()
            if val:
                config[key] = val

        missing = [k for k in ("userId", "xiaomiiot_ph", "serviceToken", "pdId", "model")
                   if not config.get(k)]
        if missing:
            QMessageBox.warning(self, "配置缺失", f"缺少必填项:\n{', '.join(missing)}")
            return None, None
        if not rows:
            QMessageBox.warning(self, "提示", "服务列表为空")
            return None, None

        # 自动注入 groupId（从当前登录用户）
        _inject_group_id(config)

        return config, rows

    def _start_dry(self):  self._run(dry_run=True)
    def _start_create(self): self._run(dry_run=False)

    def _run(self, dry_run):
        config, rows = self._build_config()
        if not config:
            return

        # ─── 优先检查产品状态（仅正式创建时）─────────────────────
        if not dry_run:
            self.log.clear()
            self.log.append("🔍 正在检查产品状态...")
            try:
                is_ok, status, status_name, msg = check_product_status(config)
                if is_ok:
                    self.log.append(f"✅ {msg}")
                else:
                    self.log.append(f"❌ {msg}")
                    QMessageBox.critical(self, "产品状态检查失败", msg)
                    return
            except Exception as e:
                self.log.append(f"⚠️ 产品状态检查异常: {e}（继续执行）")

            reply = QMessageBox.question(
                self, "确认创建", f"即将同步 {len(rows)} 个服务，是否继续？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply != QMessageBox.StandardButton.Yes:
                return
        else:
            self.log.clear()

        self.log.append(f"{'🧪 干跑模式' if dry_run else '🚀 正式创建'} - {len(rows)} 个服务\n")
        self._set_btns(running=True)
        self.progress.setVisible(True); self.progress.setRange(0, 0)

        self._worker = SyncServiceWorker(config, rows, dry_run,
                                          self.delay_spin.value() / 1000.0)
        self._worker.progress.connect(self.log.append)
        self._worker.finished_ok.connect(self._done_ok)
        self._worker.finished_err.connect(self._done_err)
        self._worker.start()

    def _cancel(self):
        if self._worker:
            self._worker.cancel()
        self.log.append("⚠️ 取消请求已发送")

    def _done_ok(self, res):
        self._set_btns(running=False)
        summary = f"创建:{res['created']} 跳过:{res['skipped']} 修正:{res['fixed']} 错误:{res['errors']}"
        self.log.append(f"\n📊 {summary}")
        if res["errors"]:
            QMessageBox.warning(self, "完成（有错误）", summary)
        else:
            QMessageBox.information(self, "完成", f"🎉 {summary}")

    def _done_err(self, msg):
        self._set_btns(running=False)
        self.log.append(f"\n❌ {msg}")
        QMessageBox.critical(self, "失败", msg)

    def _set_btns(self, running):
        self.btn_dry.setEnabled(not running)
        self.btn_run.setEnabled(not running)
        self.btn_cancel.setEnabled(running)
        self.progress.setVisible(running)


# ─── Tab: 导出服务 ────────────────────────────────────────────

class ExportServiceTab(QWidget):
    def __init__(self):
        super().__init__()
        self._worker = None
        self._build()

    def _build(self):
        layout = QHBoxLayout(self)

        left = QWidget(); left.setFixedWidth(460)
        lv = QVBoxLayout(left)

        # 连接信息
        grp_conn = QGroupBox("连接信息")
        form_conn = QFormLayout()
        self.token_edit = QLineEdit(); self.token_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.token_edit.setPlaceholderText("serviceToken")
        self.ph_edit = QLineEdit(); self.ph_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.ph_edit.setPlaceholderText("xiaomiiot_ph")
        self.userid_edit = QLineEdit(); self.userid_edit.setPlaceholderText("userId")
        self.pid_edit = QLineEdit(); self.pid_edit.setPlaceholderText("pdId（可选）")
        self.model_edit = QLineEdit(); self.model_edit.setPlaceholderText("如 uwize.switch.aiswi")
        chk = QCheckBox("显示 Cookie")
        def toggle(c):
            mode = QLineEdit.EchoMode.Normal if c else QLineEdit.EchoMode.Password
            self.token_edit.setEchoMode(mode); self.ph_edit.setEchoMode(mode)
        chk.toggled.connect(toggle)
        form_conn.addRow("serviceToken:", self.token_edit)
        form_conn.addRow("xiaomiiot_ph:", self.ph_edit)
        form_conn.addRow("userId:", self.userid_edit)
        form_conn.addRow("pdId:", self.pid_edit)
        form_conn.addRow("产品型号 (model):", self.model_edit)
        form_conn.addRow("", chk)
        grp_conn.setLayout(form_conn)
        lv.addWidget(grp_conn)

        # 或从 Excel 读取
        grp_excel = QGroupBox("或从 Excel 读取配置")
        ev = QHBoxLayout()
        self.excel_edit = QLineEdit(); self.excel_edit.setPlaceholderText("选择服务 Excel 文件")
        btn_xl = QPushButton("浏览...")
        btn_xl.clicked.connect(self._browse_excel)
        ev.addWidget(self.excel_edit); ev.addWidget(btn_xl)
        grp_excel.setLayout(ev)
        lv.addWidget(grp_excel)

        # 选项
        grp_opt = QGroupBox("导出选项")
        ov = QFormLayout()
        self.chk_props = QCheckBox("同时导出属性/事件/动作详情")
        self.out_edit = QLineEdit(); self.out_edit.setPlaceholderText("点击浏览选择导出文件夹")
        self.out_edit.setReadOnly(True)
        btn_br = QPushButton("浏览...")
        btn_br.clicked.connect(self._browse_out)
        row = QHBoxLayout(); row.addWidget(self.out_edit); row.addWidget(btn_br)
        ov.addRow("", self.chk_props)
        ov.addRow("导出文件夹:", row)
        grp_opt.setLayout(ov)
        lv.addWidget(grp_opt)

        # 按钮
        btn_row = QHBoxLayout()
        self.btn_export = QPushButton("📤 导出服务")
        self.btn_export.setObjectName("successBtn")
        self.btn_export.clicked.connect(self._start)
        btn_row.addWidget(self.btn_export)
        lv.addLayout(btn_row)
        lv.addStretch()

        # 右侧日志
        right = QWidget(); rv = QVBoxLayout(right)
        self.log = _make_log_panel(rv)
        self.progress = _make_progress(rv)

        layout.addWidget(left)
        layout.addWidget(right, stretch=1)

    def _browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择服务 Excel", "", "Excel (*.xlsx *.xls)")
        if path:
            self.excel_edit.setText(path)

    def _browse_out(self):
        path = QFileDialog.getExistingDirectory(self, "选择导出文件夹")
        if path:
            self.out_edit.setText(path)

    def _build_config(self):
        config = {
            "serviceToken": self.token_edit.text().strip(),
            "xiaomiiot_ph": self.ph_edit.text().strip(),
            "userId":       self.userid_edit.text().strip(),
            "pdId":         self.pid_edit.text().strip(),
            "model":        self.model_edit.text().strip(),
        }
        # 从 Excel 补全
        excel_path = self.excel_edit.text().strip()
        if excel_path and os.path.exists(excel_path):
            try:
                cfg_xl = read_service_config_excel(excel_path)
                for k in ("serviceToken", "xiaomiiot_ph", "userId", "pdId", "model"):
                    if not config[k]:
                        config[k] = cfg_xl.get(k, "")
            except Exception:
                pass
        missing = [k for k in ("userId", "xiaomiiot_ph", "model") if not config.get(k)]
        if missing:
            QMessageBox.warning(self, "提示", f"缺少必填项:\n{', '.join(missing)}")
            return None

        # 自动注入 groupId
        _inject_group_id(config)

        return config

    def _start(self):
        config = self._build_config()
        if not config:
            return

        out_dir = self.out_edit.text().strip()
        if not out_dir:
            out_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        safe_model = config["model"].replace(".", "_").replace("-", "_")
        output_path = os.path.join(out_dir, f"{safe_model}_services_export.xlsx")

        self.log.clear()
        self.btn_export.setEnabled(False)
        self.progress.setVisible(True); self.progress.setRange(0, 0)

        self._worker = ExportServiceWorker(
            config, output_path, export_props=self.chk_props.isChecked())
        self._worker.progress.connect(self.log.append)
        self._worker.finished_ok.connect(self._done_ok)
        self._worker.finished_err.connect(self._done_err)
        self._worker.start()

    def _done_ok(self, path):
        self.btn_export.setEnabled(True)
        self.progress.setVisible(False)
        self.log.append(f"\n🎉 导出成功: {path}")
        QMessageBox.information(self, "导出成功", f"文件已保存:\n{path}")

    def _done_err(self, msg):
        self.btn_export.setEnabled(True)
        self.progress.setVisible(False)
        self.log.append(f"\n❌ {msg}")
        QMessageBox.critical(self, "导出失败", msg)


# ─── Tab: 导出属性模板 ────────────────────────────────────────

class ExportPropTab(QWidget):
    def __init__(self):
        super().__init__()
        self._worker = None
        self._build()

    def _build(self):
        layout = QHBoxLayout(self)
        left = QWidget(); left.setFixedWidth(460)
        lv = QVBoxLayout(left)

        # 产品信息
        grp_prod = QGroupBox("产品信息")
        form = QFormLayout()
        self.pid = QLineEdit(); self.pid.setPlaceholderText("如 33257")
        self.model = QLineEdit(); self.model.setPlaceholderText("如 uwize.switch.yzw07")
        self.userid = QLineEdit(); self.userid.setPlaceholderText("如 1097752639")
        self.connect_type = QSpinBox()
        self.connect_type.setRange(0, 99); self.connect_type.setValue(16)
        form.addRow("产品ID (pdId):", self.pid)
        form.addRow("产品型号 (model):", self.model)
        form.addRow("用户ID (userId):", self.userid)
        form.addRow("连接类型:", self.connect_type)
        grp_prod.setLayout(form)
        lv.addWidget(grp_prod)

        _, self.token, self.ph, _ = _cookie_group(lv, "exp_prop", show_userid=False)

        # 输出
        grp_out = QGroupBox("输出选项")
        form2 = QFormLayout()
        self.out_edit = QLineEdit(); self.out_edit.setPlaceholderText("点击浏览选择导出文件夹")
        self.out_edit.setReadOnly(True)
        btn_br = QPushButton("浏览...")
        btn_br.clicked.connect(self._browse_out)
        row = QHBoxLayout(); row.addWidget(self.out_edit); row.addWidget(btn_br)
        form2.addRow("导出文件夹:", row)
        self.chk_json = QCheckBox("同时保存原始 JSON")
        form2.addRow("", self.chk_json)
        grp_out.setLayout(form2)
        lv.addWidget(grp_out)

        btn_row = QHBoxLayout()
        self.btn_start = QPushButton("🚀 开始导出")
        self.btn_start.setObjectName("successBtn")
        self.btn_start.clicked.connect(self._start)
        self.btn_cancel = QPushButton("取消")
        self.btn_cancel.clicked.connect(self._cancel)
        self.btn_cancel.setEnabled(False)
        btn_row.addWidget(self.btn_start); btn_row.addWidget(self.btn_cancel)
        lv.addLayout(btn_row)
        lv.addStretch()

        right = QWidget(); rv = QVBoxLayout(right)
        self.log = _make_log_panel(rv)
        self.progress = _make_progress(rv)

        layout.addWidget(left); layout.addWidget(right, stretch=1)

    def _browse_out(self):
        path = QFileDialog.getExistingDirectory(self, "选择导出文件夹")
        if path:
            self.out_edit.setText(path)

    def _start(self):
        pid = self.pid.text().strip()
        model = self.model.text().strip()
        token = self.token.text().strip()
        ph = self.ph.text().strip()
        userid = self.userid.text().strip()
        if not all([pid, model, token, ph, userid]):
            QMessageBox.warning(self, "提示", "请填写产品信息和 Cookie")
            return

        # 自动生成输出路径
        out_dir = self.out_edit.text().strip()
        if not out_dir:
            out_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        safe_model = model.replace(".", "_").replace("-", "_")
        output_path = os.path.join(out_dir, f"MIoT_模板_{safe_model}.xlsx")

        self.log.clear()
        self.btn_start.setEnabled(False); self.btn_cancel.setEnabled(True)
        self.progress.setVisible(True); self.progress.setRange(0, 0)

        self._worker = ExportPropWorker(
            pid, model, token, ph, userid,
            self.connect_type.value(),
            output_path,
            self.chk_json.isChecked(),
            0,  # 导出不需要间隔
        )
        self._worker.progress.connect(self.log.append)
        self._worker.finished_ok.connect(self._done_ok)
        self._worker.finished_err.connect(self._done_err)
        self._worker.start()

    def _cancel(self):
        if self._worker and self._worker.isRunning():
            self._worker.cancel()
            self.log.append("⚠️ 取消请求已发送")
        self._reset()

    def _done_ok(self, path):
        self._reset()
        self.log.append(f"\n🎉 导出成功: {path}")
        QMessageBox.information(self, "导出成功", f"模板已保存:\n{path}")

    def _done_err(self, msg):
        self._reset()
        self.log.append(f"\n❌ {msg}")
        QMessageBox.critical(self, "导出失败", msg)

    def _reset(self):
        self.btn_start.setEnabled(True); self.btn_cancel.setEnabled(False)
        self.progress.setVisible(False)


# ─── Tab: 创建属性 ────────────────────────────────────────────

class CreatePropTab(QWidget):
    def __init__(self):
        super().__init__()
        self._worker = None
        self._build()

    def _build(self):
        layout = QHBoxLayout(self)
        left = QWidget(); left.setFixedWidth(460)
        lv = QVBoxLayout(left)

        grp_file = QGroupBox("Excel 文件")
        fv = QHBoxLayout()
        self.file_edit = QLineEdit(); self.file_edit.setPlaceholderText("选择属性 Excel 文件")
        btn_br = QPushButton("选择文件")
        btn_br.clicked.connect(self._browse_file)
        fv.addWidget(self.file_edit); fv.addWidget(btn_br)
        grp_file.setLayout(fv)
        lv.addWidget(grp_file)

        grp_ov = QGroupBox("产品信息（可覆盖 Excel 配置）")
        form_ov = QFormLayout()
        self.pid_ov = QLineEdit(); self.pid_ov.setPlaceholderText("留空使用 Excel 配置")
        self.model_ov = QLineEdit(); self.model_ov.setPlaceholderText("留空使用 Excel 配置")
        form_ov.addRow("产品ID (pdId):", self.pid_ov)
        form_ov.addRow("产品型号 (model):", self.model_ov)
        grp_ov.setLayout(form_ov)
        lv.addWidget(grp_ov)

        _, self.token_ov, self.ph_ov, self.uid_ov = _cookie_group(lv, "crt_prop")
        self.token_ov.setPlaceholderText("留空使用 Excel 配置")
        self.ph_ov.setPlaceholderText("留空使用 Excel 配置")
        self.uid_ov.setPlaceholderText("留空使用 Excel 配置")

        grp_opts = QGroupBox("选项")
        form_opts = QFormLayout()
        self.delay_spin = QSpinBox()
        self.delay_spin.setRange(100, 2000); self.delay_spin.setValue(500)
        self.delay_spin.setSingleStep(100); self.delay_spin.setSuffix(" ms")
        self.siid_spin = QSpinBox()
        self.siid_spin.setRange(0, 999); self.siid_spin.setValue(0)
        self.siid_spin.setSpecialValueText("全部")
        form_opts.addRow("请求间隔:", self.delay_spin)
        form_opts.addRow("指定 siid:", self.siid_spin)
        grp_opts.setLayout(form_opts)
        lv.addWidget(grp_opts)

        btn_row1 = QHBoxLayout()
        self.btn_dry = QPushButton("🧪 干跑检查")
        self.btn_dry.clicked.connect(self._dryrun)
        self.btn_list = QPushButton("📋 查看服务")
        self.btn_list.clicked.connect(self._list_services)
        btn_row1.addWidget(self.btn_dry); btn_row1.addWidget(self.btn_list)
        lv.addLayout(btn_row1)

        btn_row2 = QHBoxLayout()
        self.btn_create = QPushButton("🚀 开始创建")
        self.btn_create.setObjectName("dangerBtn")
        self.btn_create.clicked.connect(self._start_create)
        self.btn_cancel = QPushButton("取消")
        self.btn_cancel.clicked.connect(self._cancel)
        self.btn_cancel.setEnabled(False)
        btn_row2.addWidget(self.btn_create); btn_row2.addWidget(self.btn_cancel)
        lv.addLayout(btn_row2)
        lv.addStretch()

        right = QWidget(); rv = QVBoxLayout(right)
        self.log = _make_log_panel(rv)
        self.progress = _make_progress(rv)

        layout.addWidget(left); layout.addWidget(right, stretch=1)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择属性 Excel", "", "Excel (*.xlsx *.xls)")
        if path:
            self.file_edit.setText(path)

    def _load(self):
        path = self.file_edit.text().strip()
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "提示", "请选择有效的 Excel 文件")
            return None, None
        from openpyxl import load_workbook
        wb = load_workbook(path)
        config = read_config(wb["公共配置"])
        props   = read_properties(wb["属性定义"]) if "属性定义" in wb.sheetnames else []
        actions = read_actions(wb["方法定义"]) if "方法定义" in wb.sheetnames else []
        events  = read_events(wb["事件定义"]) if "事件定义" in wb.sheetnames else []

        for key, widget in [
            ("serviceToken", self.token_ov), ("xiaomiiot_ph", self.ph_ov),
            ("userId", self.uid_ov), ("pdId", self.pid_ov), ("model", self.model_ov),
        ]:
            if widget.text().strip():
                config[key] = widget.text().strip()

        missing = [k for k in ("serviceToken", "xiaomiiot_ph", "userId", "pdId", "model")
                   if not config.get(k)]
        if missing:
            QMessageBox.warning(self, "配置缺失", f"缺少必填项:\n{', '.join(missing)}")
            return None, None
        if not props and not actions and not events:
            QMessageBox.warning(self, "提示", "属性/方法/事件定义均为空")
            return None, None

        # 自动注入 groupId
        _inject_group_id(config)

        return config, (props, actions, events)

    def _list_services(self):
        config, _ = self._load()
        if not config:
            return
        self.log.clear(); self.log.append("📋 查询服务列表...")
        try:
            services = create_query_services(config)
            if not services:
                self.log.append("❌ 未查到服务"); return
            self.log.append(f"\n📋 共 {len(services)} 个服务:")
            self.log.append("-" * 60)
            for svc in services:
                self.log.append(
                    f"{svc.get('siid','?'):>4} | {svc.get('name','?'):<24} | "
                    f"{svc.get('description',''):<20} | {svc.get('type','')}")
        except Exception as e:
            self.log.append(f"❌ {e}")

    def _dryrun(self):
        config, items = self._load()
        if not config:
            return
        props, actions, events = items
        self.log.clear(); self.log.append("🧪 干跑模式...\n")
        try:
            services = create_query_services(config)
            target_siid = self.siid_spin.value()

            # 属性任务
            self.log.append("📝 属性定义:")
            prop_tasks = []
            for i, prop in enumerate(props):
                svc = match_service(prop, services)
                siid = svc["siid"] if svc else prop.get("siid", "?")
                if target_siid > 0 and str(siid) != str(target_siid):
                    continue
                sname = svc.get("description", svc.get("name", "")) if svc else "❌ 未匹配"
                vtype = detect_value_type(str(prop.get("format", "")), prop)
                prop_tasks.append((i+1, prop.get("name","?"), prop.get("format","?"), vtype, siid, sname))

            if prop_tasks:
                self.log.append(f"{'#':>3} | {'name':<20} | {'format':<8} | {'vtype':<10} | siid | 服务")
                self.log.append("-" * 80)
                for t in prop_tasks:
                    self.log.append(f"{t[0]:>3} | {t[1]:<20} | {t[2]:<8} | {t[3]:<10} | {str(t[4]):<4} | {t[5]}")
            else:
                self.log.append("  （无属性）")

            # 方法任务
            self.log.append(f"\n📝 方法定义:")
            action_tasks = []
            for i, item in enumerate(actions):
                svc = match_service(item, services)
                siid = svc["siid"] if svc else item.get("siid", "?")
                if target_siid > 0 and str(siid) != str(target_siid):
                    continue
                sname = svc.get("description", svc.get("name", "")) if svc else "❌ 未匹配"
                action_tasks.append((i+1, item.get("name","?"), siid, sname))

            if action_tasks:
                self.log.append(f"{'#':>3} | {'name':<20} | siid | 服务")
                self.log.append("-" * 60)
                for t in action_tasks:
                    self.log.append(f"{t[0]:>3} | {t[1]:<20} | {str(t[2]):<4} | {t[3]}")
            else:
                self.log.append("  （无方法）")

            # 事件任务
            self.log.append(f"\n📝 事件定义:")
            event_tasks = []
            for i, item in enumerate(events):
                svc = match_service(item, services)
                siid = svc["siid"] if svc else item.get("siid", "?")
                if target_siid > 0 and str(siid) != str(target_siid):
                    continue
                sname = svc.get("description", svc.get("name", "")) if svc else "❌ 未匹配"
                event_tasks.append((i+1, item.get("name","?"), siid, sname))

            if event_tasks:
                self.log.append(f"{'#':>3} | {'name':<20} | siid | 服务")
                self.log.append("-" * 60)
                for t in event_tasks:
                    self.log.append(f"{t[0]:>3} | {t[1]:<20} | {str(t[2]):<4} | {t[3]}")
            else:
                self.log.append("  （无事件）")

            total = len(prop_tasks) + len(action_tasks) + len(event_tasks)
            self.log.append(f"\n🏁 共 {total} 条（属性{len(prop_tasks)}+方法{len(action_tasks)}+事件{len(event_tasks)}，干跑未执行）")
        except Exception:
            self.log.append(f"❌ {traceback.format_exc()}")

    def _start_create(self):
        config, items = self._load()
        if not config:
            return
        props, actions, events = items

        # ─── 优先检查产品状态 ──────────────────────────────────
        self.log.clear()
        self.log.append("🔍 正在检查产品状态...")
        try:
            is_ok, status, status_name, msg = check_product_status(config)
            if is_ok:
                self.log.append(f"✅ {msg}")
            else:
                self.log.append(f"❌ {msg}")
                QMessageBox.critical(self, "产品状态检查失败", msg)
                return
        except Exception as e:
            self.log.append(f"⚠️ 产品状态检查异常: {e}（继续执行）")

        total = len(props) + len(actions) + len(events)
        reply = QMessageBox.question(
            self, "确认创建",
            f"即将创建 {total} 条（属性{len(props)}+方法{len(actions)}+事件{len(events)}），是否继续？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            services = create_query_services(config)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"查询服务失败: {e}"); return

        target_siid = self.siid_spin.value()

        # 构建 task 列表：(type_label, item, build_fn, create_fn, id_field)
        all_tasks = []
        for p in props:
            svc = match_service(p, services)
            siid = svc["siid"] if svc else p.get("siid", "?")
            if target_siid > 0 and str(siid) != str(target_siid):
                continue
            all_tasks.append(("属性", p, build_request_body, create_property, "piid", svc))
        for a in actions:
            svc = match_service(a, services)
            siid = svc["siid"] if svc else a.get("siid", "?")
            if target_siid > 0 and str(siid) != str(target_siid):
                continue
            all_tasks.append(("方法", a, build_action_request_body, create_action, "aiid", svc))
        for e in events:
            svc = match_service(e, services)
            siid = svc["siid"] if svc else e.get("siid", "?")
            if target_siid > 0 and str(siid) != str(target_siid):
                continue
            all_tasks.append(("事件", e, build_event_request_body, create_event, "eiid", svc))

        if not all_tasks:
            QMessageBox.information(self, "提示", "没有匹配的任务"); return

        self.log.clear()
        self.log.append(f"🚀 开始创建 {len(all_tasks)} 条（属性+方法+事件）...\n")
        self._set_btns(running=True)
        self.progress.setVisible(True); self.progress.setRange(0, len(all_tasks))

        self._worker = CreateAllWorker(config, all_tasks, services, self.delay_spin.value() / 1000.0)
        self._worker.progress.connect(self.log.append)
        self._worker.update_progress.connect(lambda c, t: self.progress.setValue(c))
        self._worker.finished_ok.connect(self._done_ok)
        self._worker.finished_err.connect(self._done_err)
        self._worker.start()

    def _cancel(self):
        if self._worker and self._worker.isRunning():
            self._worker.cancel()
            self.log.append("⚠️ 取消请求已发送")
        self._set_btns(running=False)

    def _done_ok(self, success, failed):
        self._set_btns(running=False)
        self.log.append(f"\n{'='*50}\n📊 成功 {success} / 失败 {failed} / 共 {success+failed}")
        if failed:
            QMessageBox.warning(self, "完成", f"成功 {success}, 失败 {failed}")
        else:
            QMessageBox.information(self, "完成", f"🎉 全部 {success} 条创建成功！")

    def _done_err(self, msg):
        self._set_btns(running=False)
        self.log.append(f"\n❌ {msg}")
        QMessageBox.critical(self, "失败", msg)

    def _set_btns(self, running):
        for b in (self.btn_create, self.btn_dry, self.btn_list):
            b.setEnabled(not running)
        self.btn_cancel.setEnabled(running)
        self.progress.setVisible(running)


# ─── Tab: 生成模板 ────────────────────────────────────────────

class TemplatePropTab(QWidget):
    def __init__(self):
        super().__init__()
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        grp = QGroupBox("生成空白属性 Excel 模板")
        form = QFormLayout()
        self.out_edit = QLineEdit("MIoT_属性创建模板.xlsx")
        btn_br = QPushButton("浏览...")
        btn_br.clicked.connect(self._browse)
        row = QHBoxLayout(); row.addWidget(self.out_edit); row.addWidget(btn_br)
        form.addRow("输出路径:", row)
        btn_gen = QPushButton("📄 生成模板")
        btn_gen.setObjectName("successBtn")
        btn_gen.clicked.connect(self._gen)
        form.addRow("", btn_gen)
        grp.setLayout(form)
        layout.addWidget(grp)
        layout.addStretch()

    def _browse(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "选择输出路径", "MIoT_属性创建模板.xlsx", "Excel (*.xlsx)")
        if path:
            self.out_edit.setText(path)

    def _gen(self):
        path = self.out_edit.text().strip()
        if not path:
            QMessageBox.warning(self, "提示", "请填写输出路径"); return
        try:
            _generate_blank_template(path)
            QMessageBox.information(self, "成功", f"模板已生成:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "失败", str(e))


def _generate_blank_template(output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", fgColor="4472C4")
    opt_fill      = PatternFill("solid", fgColor="8DB4E2")
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border   = Border(left=Side(style="thin"), right=Side(style="thin"),
                           top=Side(style="thin"),  bottom=Side(style="thin"))
    desc_font     = Font(name="Arial", size=9, color="666666")
    desc_fill     = PatternFill("solid", fgColor="D9E2F3")
    opt_desc_fill = PatternFill("solid", fgColor="E8F0FE")

    ws = wb.active; ws.title = "属性定义"
    columns = [
        ("name",             20, "属性英文名\n如 on、mode",                               True),
        ("description",      20, "属性中文描述\n如 开关、模式",                            True),
        ("format",           12, "数据格式\nbool/uint8/uint16/uint32/string",             True),
        ("service_desc",     22, "服务中文名\n如「开关一键」",                              True),
        ("value_list",       28, "枚举值\n格式: 0:关闭,1:开启",                            False),
        ("value_range_min",  14, "数值范围-最小值",                                        False),
        ("value_range_max",  14, "数值范围-最大值",                                        False),
        ("value_range_step", 14, "数值范围-步长",                                          False),
        ("siid",              8, "服务ID（备选）",                                         False),
        ("access",           20, "访问权限\n默认: read,write,notify",                     False),
        ("service_name",     20, "服务英文名（可选）",                                      False),
    ]
    for i, (col, width, desc, required) in enumerate(columns, 1):
        cl = chr(64 + i)
        ws.column_dimensions[cl].width = width
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = header_font
        cell.fill = header_fill if required else opt_fill
        cell.alignment = header_align; cell.border = thin_border
        dc = ws.cell(row=2, column=i, value=desc)
        dc.font = desc_font
        dc.fill = desc_fill if required else opt_desc_fill
        dc.alignment = Alignment(vertical="center", wrap_text=True)
        dc.border = thin_border
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 50
    dv = DataValidation(type="list", formula1='"bool,uint8,uint16,uint32,string"', allow_blank=True)
    ws.add_data_validation(dv); dv.add("C3:C1000")

    ws2 = wb.create_sheet("公共配置")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 65
    ws2.column_dimensions["C"].width = 40
    config_items = [
        ("userId",       "", "小米账号用户ID（必填）",                    True),
        ("pdId",         "", "产品ID（必填）",                            True),
        ("model",        "", "设备型号（必填）",                           True),
        ("serviceToken", "", "浏览器 Cookie 获取（必填）",                 True),
        ("xiaomiiot_ph", "", "浏览器 Cookie 获取（必填）",                 True),
        ("connectType",  "16", "连接类型（默认16）",                       False),
        ("language",     "zh_cn", "语言（默认zh_cn）",                    False),
        ("version",      "1", "版本（默认1）",                            False),
        ("status",       "0", "状态（默认0）",                            False),
        ("source",       "4", "来源（默认4）",                            False),
        ("standard",     "false", "标准属性（默认false）",                 False),
        ("access",       "read,write,notify", "默认访问权限",              False),
    ]
    for i, (k, v, d, req) in enumerate(config_items, 1):
        kc = ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
        dc = ws2.cell(row=i, column=3, value=d)
        if req:
            kc.font = Font(name="Arial", bold=True, color="CC0000")
        dc.font = desc_font

    ws3 = wb.create_sheet("填写说明")
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 80
    instructions = [
        ("必填列", "name / description / format / service_desc"),
        ("枚举属性", "value_list 列填写格式: 0:关闭,1:开启,2:待机"),
        ("数值属性", "value_range_min / max / step 三列"),
        ("bool 属性", "format 填 bool，value_list 和 value_range 都留空"),
        ("服务匹配", "优先用 service_desc（服务中文名）匹配"),
        ("siid 列", "可选，填了则忽略 service 匹配"),
        ("access 列", "可选，默认 read,write,notify"),
    ]
    ws3.cell(row=1, column=1, value="项目").font = Font(bold=True, size=12)
    ws3.cell(row=1, column=2, value="说明").font = Font(bold=True, size=12)
    for i, (item, desc) in enumerate(instructions, 2):
        ws3.cell(row=i, column=1, value=item).font = Font(bold=True)
        ws3.cell(row=i, column=2, value=desc)

    wb.save(output_path)


# ─── 自动化 Tab ────────────────────────────────────────────────

class ExportAutomationTab(QWidget):
    """导出自定义自动化列表"""
    def __init__(self):
        super().__init__()
        self._worker = None
        self._build()

    def _build(self):
        layout = QHBoxLayout(self)

        left = QWidget(); left.setFixedWidth(460)
        lv = QVBoxLayout(left)

        # Excel 文件
        grp_excel = QGroupBox("自动化 Excel（可选，用于补全配置）")
        ev = QHBoxLayout()
        self.excel_edit = QLineEdit(); self.excel_edit.setPlaceholderText("选择已有的自动化 Excel，自动读取配置")
        btn_xl = QPushButton("浏览...")
        btn_xl.clicked.connect(self._browse_excel)
        ev.addWidget(self.excel_edit); ev.addWidget(btn_xl)
        grp_excel.setLayout(ev)
        lv.addWidget(grp_excel)

        # 产品信息覆盖
        grp_prod = QGroupBox("产品信息（可覆盖 Excel 配置）")
        form_prod = QFormLayout()
        self.pid_edit = QLineEdit(); self.pid_edit.setPlaceholderText("留空使用 Excel 配置")
        self.model_edit = QLineEdit(); self.model_edit.setPlaceholderText("留空使用 Excel 配置")
        form_prod.addRow("产品ID (pdId):", self.pid_edit)
        form_prod.addRow("产品型号 (model):", self.model_edit)
        grp_prod.setLayout(form_prod)
        lv.addWidget(grp_prod)

        # Cookie 信息（自动填充 + 手动覆盖）
        _, self.token_edit, self.ph_edit, self.userid_edit = _cookie_group(lv, "auto_exp")
        self.token_edit.setPlaceholderText("留空使用 Excel 配置或已登录账号")
        self.ph_edit.setPlaceholderText("留空使用 Excel 配置或已登录账号")
        self.userid_edit.setPlaceholderText("留空使用 Excel 配置或已登录账号")

        # 导出文件夹
        grp_out = QGroupBox("导出选项")
        ov = QFormLayout()
        self.out_edit = QLineEdit(); self.out_edit.setPlaceholderText("点击浏览选择导出文件夹")
        self.out_edit.setReadOnly(True)
        btn_br = QPushButton("浏览...")
        btn_br.clicked.connect(self._browse_out)
        row = QHBoxLayout(); row.addWidget(self.out_edit); row.addWidget(btn_br)
        ov.addRow("导出文件夹:", row)
        grp_out.setLayout(ov)
        lv.addWidget(grp_out)

        # 按钮
        btn_row = QHBoxLayout()
        self.btn_export = QPushButton("📤 导出自动化")
        self.btn_export.setObjectName("successBtn")
        self.btn_export.clicked.connect(self._start)
        btn_row.addWidget(self.btn_export)
        lv.addLayout(btn_row)
        lv.addStretch()

        # 右侧日志
        right = QWidget(); rv = QVBoxLayout(right)
        self.log = _make_log_panel(rv)
        self.progress = _make_progress(rv)

        layout.addWidget(left)
        layout.addWidget(right, stretch=1)

    def _browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择自动化 Excel", "", "Excel (*.xlsx *.xls)")
        if path:
            self.excel_edit.setText(path)

    def _browse_out(self):
        path = QFileDialog.getExistingDirectory(self, "选择导出文件夹")
        if path:
            self.out_edit.setText(path)

    def _build_config(self):
        config = {
            "serviceToken": self.token_edit.text().strip(),
            "xiaomiiot_ph": self.ph_edit.text().strip(),
            "userId":       self.userid_edit.text().strip(),
            "pdId":         self.pid_edit.text().strip(),
            "model":        self.model_edit.text().strip(),
        }
        excel_path = self.excel_edit.text().strip()
        if excel_path and os.path.exists(excel_path):
            try:
                cfg_xl, _ = read_automation_excel(excel_path)
                for k in ("serviceToken", "xiaomiiot_ph", "userId", "pdId", "model"):
                    if not config[k]:
                        config[k] = cfg_xl.get(k, "")
            except Exception:
                pass
        missing = [k for k in ("userId", "xiaomiiot_ph", "pdId") if not config.get(k)]
        if missing:
            QMessageBox.warning(self, "提示", f"缺少必填项:\n{', '.join(missing)}")
            return None

        # 自动注入 groupId
        _inject_group_id(config)

        return config

    def _start(self):
        config = self._build_config()
        if not config:
            return

        out_dir = self.out_edit.text().strip()
        if not out_dir:
            out_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        safe_model = config.get("model", "unknown").replace(".", "_").replace("-", "_")
        output_path = os.path.join(out_dir, f"{safe_model}_automation_export.xlsx")

        self.log.clear()
        self.btn_export.setEnabled(False)
        self.progress.setVisible(True); self.progress.setRange(0, 0)

        self._worker = ExportAutomationWorker(config, output_path)
        self._worker.progress.connect(self.log.append)
        self._worker.finished_ok.connect(self._done_ok)
        self._worker.finished_err.connect(self._done_err)
        self._worker.start()

    def _done_ok(self, path):
        self.btn_export.setEnabled(True)
        self.progress.setVisible(False)
        self.log.append(f"\n🎉 导出成功: {path}")
        QMessageBox.information(self, "导出成功", f"文件已保存:\n{path}")

    def _done_err(self, msg):
        self.btn_export.setEnabled(True)
        self.progress.setVisible(False)
        self.log.append(f"\n❌ {msg}")
        QMessageBox.critical(self, "导出失败", msg)


class CreateAutomationTab(QWidget):
    """批量创建自定义自动化"""
    def __init__(self):
        super().__init__()
        self._worker = None
        self._build()

    def _build(self):
        layout = QHBoxLayout(self)

        left = QWidget(); left.setFixedWidth(460)
        lv = QVBoxLayout(left)

        # Excel 文件
        grp_excel = QGroupBox("自动化 Excel（包含配置 + 自动化列表）")
        ev = QHBoxLayout()
        self.excel_edit = QLineEdit(); self.excel_edit.setPlaceholderText("选择自动化 Excel 文件")
        btn_xl = QPushButton("浏览...")
        btn_xl.clicked.connect(self._browse_file)
        ev.addWidget(self.excel_edit); ev.addWidget(btn_xl)
        grp_excel.setLayout(ev)
        lv.addWidget(grp_excel)

        # 产品信息覆盖
        grp_prod = QGroupBox("产品信息（可覆盖 Excel 配置）")
        form_prod = QFormLayout()
        self.pid_edit = QLineEdit(); self.pid_edit.setPlaceholderText("留空使用 Excel 配置")
        self.model_edit = QLineEdit(); self.model_edit.setPlaceholderText("留空使用 Excel 配置")
        form_prod.addRow("产品ID (pdId):", self.pid_edit)
        form_prod.addRow("产品型号 (model):", self.model_edit)
        grp_prod.setLayout(form_prod)
        lv.addWidget(grp_prod)

        # Cookie 信息（自动填充 + 手动覆盖）
        _, self.token_edit, self.ph_edit, self.userid_edit = _cookie_group(lv, "auto_crt")
        self.token_edit.setPlaceholderText("留空使用 Excel 配置或已登录账号")
        self.ph_edit.setPlaceholderText("留空使用 Excel 配置或已登录账号")
        self.userid_edit.setPlaceholderText("留空使用 Excel 配置或已登录账号")

        # 选项
        grp_opt = QGroupBox("选项")
        ov = QFormLayout()
        self.chk_dryrun = QCheckBox("Dry-run（仅预检，不实际创建）")
        ov.addRow("", self.chk_dryrun)
        self.delay_spin = QSpinBox(); self.delay_spin.setRange(100, 2000)
        self.delay_spin.setValue(500); self.delay_spin.setSingleStep(100)
        self.delay_spin.setSuffix(" ms")
        ov.addRow("请求间隔:", self.delay_spin)
        grp_opt.setLayout(ov)
        lv.addWidget(grp_opt)

        # 按钮
        btn_row = QHBoxLayout()
        self.btn_create = QPushButton("🚀 创建自动化")
        self.btn_create.setObjectName("successBtn")
        self.btn_create.clicked.connect(self._start)
        self.btn_cancel = QPushButton("⏹ 取消")
        self.btn_cancel.setObjectName("dangerBtn")
        self.btn_cancel.clicked.connect(self._cancel)
        self.btn_cancel.setEnabled(False)
        btn_row.addWidget(self.btn_create)
        btn_row.addWidget(self.btn_cancel)
        lv.addLayout(btn_row)
        lv.addStretch()

        # 右侧日志
        right = QWidget(); rv = QVBoxLayout(right)
        self.log = _make_log_panel(rv)
        self.progress = _make_progress(rv)

        layout.addWidget(left)
        layout.addWidget(right, stretch=1)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择自动化 Excel", "", "Excel (*.xlsx *.xls)")
        if path:
            self.excel_edit.setText(path)

    def _start(self):
        excel_path = self.excel_edit.text().strip()
        if not excel_path or not os.path.exists(excel_path):
            QMessageBox.warning(self, "提示", "请先选择自动化 Excel 文件")
            return

        try:
            config, auto_items = read_automation_excel(excel_path)
        except Exception as e:
            QMessageBox.critical(self, "读取失败", f"Excel 读取错误:\n{e}")
            return

        # 手动输入覆盖 Excel 中的配置
        manual = {
            "serviceToken": self.token_edit.text().strip(),
            "xiaomiiot_ph": self.ph_edit.text().strip(),
            "userId":       self.userid_edit.text().strip(),
            "pdId":         self.pid_edit.text().strip(),
            "model":        self.model_edit.text().strip(),
        }
        for k, v in manual.items():
            if v:
                config[k] = v

        missing = [k for k in ("userId", "xiaomiiot_ph", "pdId") if not config.get(k)]
        if missing:
            QMessageBox.warning(self, "提示", f"缺少必填项:\n{', '.join(missing)}")
            return

        # 自动注入 groupId
        _inject_group_id(config)

        if not auto_items:
            QMessageBox.warning(self, "提示", "自动化列表为空")
            return

        # ─── 优先检查产品状态 ──────────────────────────────────
        self.log.clear()
        self.log.append("🔍 正在检查产品状态...")
        try:
            is_ok, status, status_name, msg = check_product_status(config)
            if is_ok:
                self.log.append(f"✅ {msg}")
            else:
                self.log.append(f"❌ {msg}")
                QMessageBox.critical(self, "产品状态检查失败", msg)
                return
        except Exception as e:
            self.log.append(f"⚠️ 产品状态检查异常: {e}（继续执行）")

        dry = self.chk_dryrun.isChecked()
        delay = self.delay_spin.value() / 1000.0

        self.log.append(f"📋 共 {len(auto_items)} 个自动化待创建" + (" (dry-run)" if dry else ""))
        self.btn_create.setEnabled(False)
        self.btn_cancel.setEnabled(True)
        self.progress.setVisible(True); self.progress.setRange(0, len(auto_items))

        self._worker = CreateAutomationWorker(
            config, auto_items, dry_run=dry, delay=delay)
        self._worker.progress.connect(self.log.append)
        self._worker.update_progress.connect(lambda c, t: self.progress.setValue(c))
        self._worker.finished_ok.connect(self._done_ok)
        self._worker.finished_err.connect(self._done_err)
        self._worker.start()

    def _cancel(self):
        if self._worker and self._worker.isRunning():
            self._worker.cancel()
        self.log.append("⚠️ 取消请求已发送")

    def _done_ok(self, success, failed):
        self.btn_create.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.progress.setVisible(False)
        self.log.append(f"\n🎉 完成！成功: {success}, 失败: {failed}")
        QMessageBox.information(self, "创建完成", f"成功: {success}\n失败: {failed}")

    def _done_err(self, msg):
        self.btn_create.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.progress.setVisible(False)
        self.log.append(f"\n❌ {msg}")
        QMessageBox.critical(self, "创建失败", msg)


# ─── 登录对话框 ──────────────────────────────────────────────

class LoginDialog(QDialog):
    """小米账号登录对话框 - 内嵌浏览器"""
    login_success = pyqtSignal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("登录小米账号")
        self.setMinimumSize(480, 640)
        self.resize(520, 700)
        self._browser = MiLoginBrowser(self)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # 浏览器
        view = self._browser.create_view()
        layout.addWidget(view)

        # 连接信号
        self._browser.login_success.connect(self._on_login_success)

        # 启动登录
        self._browser.start_login()

    def _on_login_success(self, user_info: dict):
        """登录成功"""
        save_user(
            user_id=user_info["userId"],
            service_token=user_info["serviceToken"],
            xiaomiiot_ph=user_info["xiaomiiot_ph"],
            name=user_info.get("userId", ""),
            group_id=user_info.get("groupId", ""),
        )
        self.login_success.emit(user_info)
        self.accept()

    def closeEvent(self, event):
        self._browser.cleanup()
        super().closeEvent(event)


# ─── 自定义企业下拉框（弹出列表自动展开宽度） ────────────────

class EnterpriseComboBox(QComboBox):
    """下拉弹出列表宽度根据内容自动展开，不受控件本身固定宽度限制"""

    def showPopup(self):
        super().showPopup()
        view = self.view()
        if view is None:
            return
        fm = self.fontMetrics()
        max_text_width = 0
        for i in range(self.count()):
            text = self.itemText(i)
            max_text_width = max(max_text_width, fm.horizontalAdvance(text))
        needed_width = max_text_width + 60
        combo_width = self.width()
        popup_width = max(needed_width, combo_width)
        popup = view.parentWidget()
        if popup:
            popup.setFixedWidth(popup_width)


# ─── Main Window ──────────────────────────────────────────────

class MIoTMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("MIoT 平台工具")
        self.setMinimumSize(1020, 820)
        self.resize(1060, 860)
        self._current_user = None
        self._ent_loading = False
        self._init_ui()
        self._update_user_ui()  # 初始化用户区域状态
        self._check_saved_login()

    def _init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(16, 12, 16, 12)

        # ── 标题行 + 用户区域
        header = QHBoxLayout()

        title = QLabel("🔧 MIoT 平台工具")
        title.setObjectName("titleLabel")
        subtitle = QLabel("小米 IoT 平台  —  服务层管理 & 属性层管理 & 自定义自动化（整合版）")
        subtitle.setObjectName("subtitleLabel")

        header_left = QVBoxLayout()
        header_left.addWidget(title)
        header_left.addWidget(subtitle)
        header_left.setSpacing(2)
        header.addLayout(header_left, 1)

        # 右上角用户区域
        self._build_user_area(header)

        layout.addLayout(header)

        # 外层 Tabs：服务层 / 属性层
        self.outer_tabs = QTabWidget()
        layout.addWidget(self.outer_tabs)

        # ── 服务层
        svc_widget = QWidget()
        svc_layout = QVBoxLayout(svc_widget)
        svc_layout.setContentsMargins(0, 0, 0, 0)
        svc_inner = QTabWidget()
        svc_inner.addTab(ExportServiceTab(), "📤 导出服务")
        svc_inner.addTab(CreateServiceTab(), "📋 创建服务")
        svc_layout.addWidget(svc_inner)
        self.outer_tabs.addTab(svc_widget, "🏗️ 服务层")

        # ── 属性层
        prop_widget = QWidget()
        prop_layout = QVBoxLayout(prop_widget)
        prop_layout.setContentsMargins(0, 0, 0, 0)
        prop_inner = QTabWidget()
        prop_inner.addTab(ExportPropTab(),   "📤 导出模板")
        prop_inner.addTab(CreatePropTab(),   "📥 创建属性")
        prop_inner.addTab(TemplatePropTab(), "📄 生成模板")
        prop_layout.addWidget(prop_inner)
        self.outer_tabs.addTab(prop_widget, "⚙️ 属性层")

        # ── 自动化
        auto_widget = QWidget()
        auto_layout = QVBoxLayout(auto_widget)
        auto_layout.setContentsMargins(0, 0, 0, 0)
        auto_inner = QTabWidget()
        auto_inner.addTab(ExportAutomationTab(), "📤 导出自动化")
        auto_inner.addTab(CreateAutomationTab(), "📥 创建自动化")
        auto_layout.addWidget(auto_inner)
        self.outer_tabs.addTab(auto_widget, "🤖 自动化")

        self.statusBar().showMessage("就绪")

    # ─── 用户区域 ─────────────────────────────────────────────

    def _build_user_area(self, parent_layout):
        """构建右上角用户区域：[企业下拉] [用户按钮]"""
        user_row = QHBoxLayout()
        user_row.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        user_row.setSpacing(8)

        # 企业下拉（左侧）
        self.ent_combo = EnterpriseComboBox()
        self.ent_combo.setObjectName("entCombo")
        self.ent_combo.setCursor(Qt.CursorShape.PointingHandCursor)
        self.ent_combo.setVisible(False)  # 登录后才显示
        self.ent_combo.setToolTip("切换当前企业")
        self.ent_combo.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon)
        self.ent_combo.setMinimumContentsLength(10)
        self.ent_combo.setFixedWidth(280)
        self.ent_combo.currentIndexChanged.connect(self._on_ent_combo_changed)
        user_row.addWidget(self.ent_combo)

        # 用户按钮（右侧）
        self.user_btn = QPushButton("🔑 点击登录")
        self.user_btn.setObjectName("userBtn")
        self.user_btn.setProperty("loggedIn", "false")
        self.user_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.user_btn.clicked.connect(self._on_user_btn_clicked)
        user_row.addWidget(self.user_btn)

        parent_layout.addLayout(user_row)

    def _on_user_btn_clicked(self):
        """点击用户按钮 - 弹出菜单"""
        if not self._current_user:
            self._open_login()
            return

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu { font-size: 13px; padding: 4px; }
            QMenu::item { padding: 6px 20px; }
            QMenu::item:selected { background-color: #eaf2f8; }
        """)

        # 切换用户
        all_users = get_all_users()
        if len(all_users) > 1:
            switch_menu = menu.addMenu("🔄 切换用户")
            for u in all_users:
                uid = u["userId"]
                is_current = (uid == str(self._current_user.get("userId", "")))
                label = f"{'✅ ' if is_current else ''}{u['name']} ({uid})"
                act = switch_menu.addAction(label)
                act.setData(uid)
                if is_current:
                    act.setEnabled(False)
            switch_menu.triggered.connect(self._on_switch_user)
            menu.addSeparator()

        # 退出登录
        menu.addAction("🚪 退出登录", self._on_logout)
        # 删除用户
        menu.addAction("🗑️ 删除此账号", self._on_delete_user)

        # 在按钮下方弹出
        pos = self.user_btn.mapToGlobal(self.user_btn.rect().bottomLeft())
        menu.exec(pos)

    def _on_ent_combo_changed(self, index):
        """企业下拉切换"""
        if index < 0 or self._ent_loading:
            return
        ent = self.ent_combo.currentData()
        if not ent or not isinstance(ent, dict):
            return
        gid = ent.get("groupId", "")
        if not gid:
            return

        cur_gid = str(self._current_user.get("groupId", ""))
        if gid == cur_gid:
            return  # 没有变化

        # 调用 API 切换企业
        self.statusBar().showMessage(f"正在切换到 {ent.get('shortName', gid)}...", 3000)
        ok = set_curr_enterprise(
            self._current_user.get("userId", ""),
            self._current_user.get("xiaomiiot_ph", ""),
            self._current_user.get("serviceToken", ""),
            gid,
            ent.get("shortName", ""),
            ent.get("compName", ""),
        )
        if ok:
            update_user_group(self._current_user.get("userId", ""), gid)
            self._current_user["groupId"] = gid
            self._current_user["groupName"] = ent.get("compName", "")
            self.statusBar().showMessage(
                f"✅ 已切换到 {ent.get('shortName', gid)} ({ent.get('compName', '')})", 5000)
        else:
            QMessageBox.warning(self, "切换失败", "切换企业失败，请重试")
            # 回滚选择
            self._ent_loading = True
            self._select_current_enterprise()
            self._ent_loading = False

    def _on_switch_user(self, action):
        """切换用户"""
        uid = action.data()
        if uid:
            user = switch_user(uid)
            if user:
                self._current_user = user
                self._update_user_ui()
                self._fill_cookies()

    def _on_logout(self):
        """退出当前用户"""
        logout_current()
        self._current_user = None
        self._update_user_ui()
        self._clear_cookies()

    def _on_delete_user(self):
        """删除当前用户"""
        if self._current_user:
            uid = str(self._current_user.get("userId", ""))
            ret = QMessageBox.question(
                self, "确认删除",
                f"确定要删除用户 {uid} 的登录信息吗？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if ret == QMessageBox.StandardButton.Yes:
                remove_user(uid)
                self._current_user = get_current_user()
                self._update_user_ui()
                if self._current_user:
                    self._fill_cookies()
                else:
                    self._clear_cookies()

    def _open_login(self):
        """打开登录对话框"""
        dlg = LoginDialog(self)
        dlg.login_success.connect(self._on_login_success)
        dlg.exec()

    def _on_login_success(self, user_info: dict):
        """登录成功回调"""
        self._current_user = user_info
        self._update_user_ui()
        self._fill_cookies()

    def _update_user_ui(self):
        """更新用户区域和企业下拉 UI"""
        if self._current_user:
            self.ent_combo.setVisible(True)
            self.ent_combo.setEnabled(True)
            uid = str(self._current_user.get("userId", ""))
            name = self._current_user.get("name", uid)
            self.user_btn.setText(f"👤 {name}")
            self.user_btn.setProperty("loggedIn", "true")
            status_msg = f"已登录: {name} ({uid})"
            self._refresh_ent_combo()
            gid = self._current_user.get("groupId", "")
            if gid:
                status_msg += f" | 企业: {gid}"
            self.statusBar().showMessage(status_msg, 5000)
        else:
            self.ent_combo.setVisible(False)
            self.ent_combo.clear()
            self.user_btn.setText("🔑 点击登录")
            self.user_btn.setProperty("loggedIn", "false")
            self.statusBar().showMessage("未登录", 3000)
        # 刷新 QSS（property 变化需要重新应用样式）
        self.user_btn.style().unpolish(self.user_btn)
        self.user_btn.style().polish(self.user_btn)

    def _refresh_ent_combo(self):
        """刷新企业下拉列表并选中当前企业"""
        if not self._current_user:
            return
        self._ent_loading = True
        self.ent_combo.blockSignals(True)
        self.ent_combo.clear()

        try:
            enterprises = get_enterprise_list(
                self._current_user.get("userId", ""),
                self._current_user.get("xiaomiiot_ph", ""),
                self._current_user.get("serviceToken", ""),
            )
            cur_gid = str(self._current_user.get("groupId", ""))
            current_index = -1
            for i, ent in enumerate(enterprises):
                gid = ent.get("groupId", "")
                # 只显示中文企业名称
                display_name = ent.get("compName", "") or ent.get("shortName", gid)
                self.ent_combo.addItem(display_name, ent)
                if gid == cur_gid:
                    current_index = i

            if current_index >= 0:
                self.ent_combo.setCurrentIndex(current_index)
            elif enterprises:
                self.ent_combo.setCurrentIndex(0)
            else:
                self.ent_combo.addItem("（无企业）")
                self.ent_combo.setEnabled(False)
        except Exception:
            self.ent_combo.addItem("（查询失败）")
            self.ent_combo.setEnabled(False)

        self.ent_combo.blockSignals(False)
        self._ent_loading = False

    def _select_current_enterprise(self):
        """根据 _current_user 中的 groupId 选中对应项（不重新请求API）"""
        cur_gid = str(self._current_user.get("groupId", ""))
        self.ent_combo.blockSignals(True)
        for i in range(self.ent_combo.count()):
            ent = self.ent_combo.itemData(i)
            if ent and isinstance(ent, dict) and str(ent.get("groupId", "")) == cur_gid:
                self.ent_combo.setCurrentIndex(i)
                break
        self.ent_combo.blockSignals(False)

    def _fill_cookies(self):
        """自动填充所有 Tab 中的 Cookie 字段"""
        if not self._current_user:
            return
        token = self._current_user.get("serviceToken", "")
        ph = self._current_user.get("xiaomiiot_ph", "")
        uid = str(self._current_user.get("userId", ""))

        # 遍历所有 Tab 中的 Cookie 字段
        for tab_widget in self._find_all_tabs():
            self._fill_tab_cookies(tab_widget, token, ph, uid)

    def _find_all_tabs(self) -> list:
        """找到所有内层 Tab 页"""
        tabs = []
        for i in range(self.outer_tabs.count()):
            outer_page = self.outer_tabs.widget(i)
            inner_tabs = outer_page.findChild(QTabWidget)
            if inner_tabs:
                for j in range(inner_tabs.count()):
                    tabs.append(inner_tabs.widget(j))
        return tabs

    def _fill_tab_cookies(self, widget, token, ph, uid):
        """填充单个 Tab 中的 Cookie 字段"""
        # 查找所有 QLineEdit，按 placeholder 或 objectName 识别
        for edit in widget.findChildren(QLineEdit):
            name = edit.placeholderText().lower()
            obj = edit.objectName().lower() if edit.objectName() else ""
            if "servicetoken" in name or "token" in obj:
                if not edit.text().strip():
                    edit.setText(token)
            elif "xiaomiiot_ph" in name or "ph" in obj:
                if not edit.text().strip():
                    edit.setText(ph)
            elif "userid" in name or "uid" in obj:
                if not edit.text().strip():
                    edit.setText(uid)
        # 也通过变量名模式匹配（更可靠）
        for attr_name in dir(widget):
            if attr_name.startswith("_"):
                continue
            attr = getattr(widget, attr_name, None)
            if not isinstance(attr, QLineEdit):
                continue
            al = attr_name.lower()
            if "token" in al and not attr.text().strip():
                attr.setText(token)
            elif "ph" in al and "edit" in al and not attr.text().strip():
                attr.setText(ph)
            elif ("userid" in al or "uid" in al) and "edit" in al and not attr.text().strip():
                attr.setText(uid)

    def _clear_cookies(self):
        """清空所有 Tab 中的 Cookie 字段"""
        for tab_widget in self._find_all_tabs():
            for edit in tab_widget.findChildren(QLineEdit):
                al = edit.objectName().lower() if edit.objectName() else ""
                attr_name = ""
                # 通过变量名模式查找
                for an in dir(tab_widget):
                    if getattr(tab_widget, an, None) is edit:
                        attr_name = an.lower()
                        break
                if any(k in attr_name for k in ("token", "ph_edit", "userid_edit", "uid_edit")):
                    edit.clear()

    def _check_saved_login(self):
        """检查是否有已保存的登录信息"""
        user = get_current_user()
        if user:
            self._current_user = user
            # 如果本地没有 groupId，尝试从 API 获取
            if not user.get("groupId"):
                try:
                    ent = get_curr_enterprise(
                        user.get("userId", ""),
                        user.get("xiaomiiot_ph", ""),
                        user.get("serviceToken", ""),
                    )
                    if ent.get("groupId"):
                        user["groupId"] = ent["groupId"]
                        user["groupName"] = ent.get("compName", "")
                        update_user_group(user.get("userId", ""), ent["groupId"])
                except Exception:
                    pass
            self._update_user_ui()
            self._fill_cookies()


# ─── Entry ────────────────────────────────────────────────────

def main():
    # WebEngine 必须在 QApplication 创建前导入
    from PyQt6.QtWebEngineWidgets import QWebEngineView  # noqa: F401

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setStyleSheet(STYLESHEET)
    window = MIoTMainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
